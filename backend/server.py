from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, APIRouter, Depends, HTTPException, Request, Response, UploadFile, File, Form, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
from datetime import datetime, timedelta, timezone, date, time
from zoneinfo import ZoneInfo
from typing import Optional, List
import os
import calendar
import bcrypt
import jwt
import logging
import uuid
import io
import traceback
import requests as http_requests
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

# ─────────────────────────────────────────────
# Constants / Config
# ─────────────────────────────────────────────
ROOT_DIR = Path(__file__).parent
MYT = ZoneInfo("Asia/Kuala_Lumpur")
JWT_SECRET = os.environ.get("JWT_SECRET", "change-this-secret")
JWT_ALGORITHM = "HS256"
LOCKOUT_ATTEMPTS = 5
LOCKOUT_MINUTES = 15
LATE_ARRIVAL_THRESHOLD_MINUTES = int(os.environ.get("LATE_ARRIVAL_THRESHOLD_MINUTES", "15"))
# Cookie settings — set COOKIE_SECURE=true on Render (HTTPS) for cross-origin cookie support
COOKIE_SECURE = os.environ.get("COOKIE_SECURE", "false").lower() == "true"
COOKIE_SAMESITE = "none" if COOKIE_SECURE else "lax"

# ─────────────────────────────────────────────
# ULTRAMSG / WHATSAPP / CRON
# ─────────────────────────────────────────────
ULTRAMSG_INSTANCE_ID = os.environ.get("ULTRAMSG_INSTANCE_ID")
ULTRAMSG_TOKEN = os.environ.get("ULTRAMSG_TOKEN")
BOSS_WHATSAPP_NUMBER = os.environ.get("BOSS_WHATSAPP_NUMBER")
CRON_SECRET = os.environ.get("CRON_SECRET")

# ─────────────────────────────────────────────
# OBJECT STORAGE
# ─────────────────────────────────────────────
STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
EMERGENT_LLM_KEY = os.environ.get("EMERGENT_LLM_KEY")
APP_STORAGE_PREFIX = "performance-pulse"
_storage_key = None

def init_storage():
    global _storage_key
    if _storage_key:
        return _storage_key
    resp = http_requests.post(f"{STORAGE_URL}/init", json={"emergent_key": EMERGENT_LLM_KEY}, timeout=30)
    resp.raise_for_status()
    _storage_key = resp.json()["storage_key"]
    return _storage_key

def put_object(path: str, data: bytes, content_type: str, timeout: int = 120) -> dict:
    key = init_storage()
    resp = http_requests.put(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key, "Content-Type": content_type},
        data=data, timeout=timeout
    )
    resp.raise_for_status()
    return resp.json()

def get_object(path: str):
    key = init_storage()
    resp = http_requests.get(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key}, timeout=60
    )
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")

# ─────────────────────────────────────────────
# WHATSAPP NOTIFIER (UltraMsg)
# ─────────────────────────────────────────────

def send_whatsapp(message: str):
    if not ULTRAMSG_INSTANCE_ID or not ULTRAMSG_TOKEN or not BOSS_WHATSAPP_NUMBER:
        raise RuntimeError("UltraMsg credentials not configured")
    url = f"https://api.ultramsg.com/{ULTRAMSG_INSTANCE_ID}/messages/chat"
    payload = {
        "token": ULTRAMSG_TOKEN,
        "to": BOSS_WHATSAPP_NUMBER,
        "body": message,
        "priority": 10
    }
    response = http_requests.post(url, json=payload, timeout=30)
    response.raise_for_status()
    return response.json()

# ─────────────────────────────────────────────
# EXCEL HELPERS
# ─────────────────────────────────────────────
REPORT_COL_HEADERS = [
    "Date", "Employee Name", "Department", "Job Title",
    "Morning Planned Tasks 9 AM – 12/1 PM", "Afternoon Planned Tasks 1/2 PM – 6 PM",
    "Final Completed Work Submit by 8 PM", "Task Category", "Task Status",
    "Calls Made", "Follow-ups", "Interested Leads", "Blockers / Issues",
    "Final Remarks", "Morning Plan Matched?", "Afternoon Plan Matched?",
    "Overall Tally Status", "Submitted By 8 PM?", "Review Status", "Manager Notes"
]

def generate_excel_template(employee_name: str = "", department: str = "") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Report"

    ws.merge_cells("A1:T1")
    ws["A1"] = "Employee Daily Reporting Template — Morning Plan, Afternoon Plan & 8 PM Final Report"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor="1E3A5F")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:T2")
    ws["A2"] = "Rule: Employees plan tasks before/after lunch and submit the final daily report by 8 PM. Final report must tally with the planned tasks."
    ws["A2"].font = Font(italic=True, size=10, color="444444")
    ws["A2"].fill = PatternFill("solid", fgColor="EFF6FF")

    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for idx, header in enumerate(REPORT_COL_HEADERS, start=1):
        cell = ws.cell(row=4, column=idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[4].height = 45

    today_str = get_myt_today()
    ws["A5"] = date.fromisoformat(today_str)
    ws["A5"].number_format = "YYYY-MM-DD"
    ws["B5"] = employee_name
    ws["C5"] = department

    col_widths = [14, 20, 15, 15, 35, 35, 35, 20, 15, 10, 10, 12, 25, 25, 18, 18, 18, 16, 15, 25]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    cat_vals = '"Admin Work,Letter Preparation,Client Calling,Digital Marketing,App Testing,Follow-up,Other"'
    dv_cat = DataValidation(type="list", formula1=cat_vals, allow_blank=True)
    dv_cat.sqref = "H5:H100"
    ws.add_data_validation(dv_cat)

    status_vals = '"Completed,In Progress,Pending,Delayed"'
    dv_status = DataValidation(type="list", formula1=status_vals, allow_blank=True)
    dv_status.sqref = "I5:I100"
    ws.add_data_validation(dv_status)

    # Add Weekly Summary sheet (placeholder)
    ws2 = wb.create_sheet("Weekly Summary")
    ws2["A1"] = "Weekly Summary — auto-generated by system"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def parse_daily_report_excel(file_bytes: bytes) -> list:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Daily Report" not in wb.sheetnames:
        raise ValueError("Excel file must contain a 'Daily Report' sheet")
    ws = wb["Daily Report"]
    rows_data = []

    for row in ws.iter_rows(min_row=5, values_only=True):
        row = list(row) + [None] * (20 - len(row))
        a, b, c, d, e, f, g, h, i, j, k, leads_val, m, n, o, p, q, r, s, t = row[:20]

        if all(v is None or str(v).strip() == "" for v in [a, b, e, f, g]):
            break
        if a is None:
            continue

        if hasattr(a, "strftime"):
            report_date = a.strftime("%Y-%m-%d")
        else:
            try:
                report_date = datetime.strptime(str(a).strip(), "%Y-%m-%d").strftime("%Y-%m-%d")
            except Exception:
                report_date = str(a).strip()

        def safe_int(v):
            try:
                return max(0, int(float(str(v)))) if v is not None and str(v).strip() else 0
            except Exception:
                return 0

        rows_data.append({
            "report_date": report_date,
            "employee_name": str(b or "").strip(),
            "department": str(c or "").strip(),
            "job_title": str(d or "").strip(),
            "morning_plan": str(e or "").strip(),
            "afternoon_plan": str(f or "").strip(),
            "final_report": str(g or "").strip(),
            "task_category": str(h or "Other").strip() or "Other",
            "task_status": str(i or "Completed").strip() or "Completed",
            "calls_made": safe_int(j),
            "follow_ups": safe_int(k),
            "interested_leads": safe_int(leads_val),
            "blockers": str(m or "").strip(),
            "final_remarks": str(n or "").strip(),
            "morning_plan_matched": str(o or "").strip(),
            "afternoon_plan_matched": str(p or "").strip(),
            "overall_tally_status": str(q or "").strip(),
            "submitted_by_6pm_excel": str(r or "").strip(),
            "review_status_from_excel": str(s or "").strip(),
            "manager_notes": str(t or "").strip(),
        })
    return rows_data

# ─────────────────────────────────────────────
# PERFORMANCE SCORE
# ─────────────────────────────────────────────

def calculate_performance_score(submitted: int, working_days: int, completed: int,
                                  delayed: int, missing: int, total_calls: int) -> dict:
    sub_rate = submitted / max(1, working_days)
    submission_score = round(min(30, sub_rate * 30))

    comp_rate = completed / max(1, submitted)
    completion_score = round(min(25, comp_rate * 25))

    avg_calls = total_calls / max(1, submitted)
    call_score = min(20, round((avg_calls / 5) * 20))

    delay_score = max(0, 15 - delayed * 3)
    missing_score = max(0, 10 - missing * 2)

    total = submission_score + completion_score + call_score + delay_score + missing_score

    if total >= 80:
        level = "Strong"
    elif total >= 60:
        level = "Good"
    elif total >= 40:
        level = "Average"
    else:
        level = "Needs Improvement"

    red_flags = []
    if missing > 2:
        red_flags.append(f"{missing} missing report days")
    if delayed > 3:
        red_flags.append(f"{delayed} delayed tasks")
    if sub_rate < 0.7 and working_days > 0:
        red_flags.append(f"Submission rate only {round(sub_rate * 100)}%")
    if avg_calls < 2 and submitted > 0:
        red_flags.append(f"Low call activity (avg {avg_calls:.1f}/day)")

    return {
        "performance_score": total,
        "performance_level": level,
        "red_flags": red_flags,
        "score_breakdown": {
            "submission": submission_score,
            "completion": completion_score,
            "calls": call_score,
            "delays": delay_score,
            "missing": missing_score,
        }
    }

# MongoDB
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# App
app = FastAPI()
api_router = APIRouter(prefix="/api")

# CORS
cors_origins_raw = os.environ.get("CORS_ORIGINS", "http://localhost:3000").split(",")
cors_origins = [o.strip() for o in cors_origins_raw if o.strip() and o.strip() != "*"]
if not cors_origins:
    cors_origins = ["http://localhost:3000"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

DEPARTMENTS = ["Sales", "Marketing", "Operations", "IT", "HR", "Management", "Other"]
TASK_CATEGORIES = ["Admin Work", "Letter Preparation", "Client Calling", "Digital Marketing", "App Testing", "Follow-up", "Other"]
TASK_STATUSES = ["Completed", "In Progress", "Pending", "Delayed"]
REVIEW_STATUSES = ["submitted", "reviewed", "needs_correction"]

# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────

def doc_to_dict(doc: dict) -> dict:
    if doc is None:
        return None
    d = dict(doc)
    if "_id" in d:
        d["id"] = str(d.pop("_id"))
    return d

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

def create_access_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(hours=8),
        "type": "access"
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "refresh"
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def set_auth_cookies(response: Response, access_token: str, refresh_token: str):
    response.set_cookie("access_token", access_token, httponly=True, secure=COOKIE_SECURE, samesite=COOKIE_SAMESITE, max_age=28800, path="/")
    response.set_cookie("refresh_token", refresh_token, httponly=True, secure=COOKIE_SECURE, samesite=COOKIE_SAMESITE, max_age=604800, path="/")

# Malaysia timezone helpers
def get_myt_today() -> str:
    return datetime.now(MYT).strftime("%Y-%m-%d")

def get_myt_now() -> datetime:
    return datetime.now(MYT)

def get_saturdays_in_month(year: int, month: int) -> List[date]:
    """Return all Saturdays in the given month."""
    saturdays = []
    cal = calendar.monthcalendar(year, month)
    for week in cal:
        if week[calendar.SATURDAY] != 0:
            saturdays.append(date(year, month, week[calendar.SATURDAY]))
    return saturdays

def is_working_day(d: date) -> bool:
    """Check if a date is a working day.
    Mon-Fri: always working
    Saturday: only first and last Saturday of the month are working (half-days)
    Sunday: never working
    """
    weekday = d.weekday()
    if weekday < 5:  # Monday-Friday
        return True
    if weekday == 6:  # Sunday
        return False
    # Saturday - check if first or last of the month
    saturdays = get_saturdays_in_month(d.year, d.month)
    if not saturdays:
        return False
    return d == saturdays[0] or d == saturdays[-1]

def is_half_day(d: date) -> bool:
    """Check if a date is a half-day (first or last Saturday of the month)."""
    if d.weekday() != 5:  # Not Saturday
        return False
    saturdays = get_saturdays_in_month(d.year, d.month)
    if not saturdays:
        return False
    return d == saturdays[0] or d == saturdays[-1]

def get_auto_clockout_time(d: date) -> Optional[time]:
    """Get the auto clock-out time for a given date.
    Mon-Fri: 18:00 (6 PM)
    First/last Saturday: 13:00 (1 PM)
    Other days: None (not a working day)
    """
    if not is_working_day(d):
        return None
    if is_half_day(d):
        return time(13, 0, 0)  # 1 PM
    return time(18, 0, 0)  # 6 PM

def get_working_days_in_range(start: date, end: date) -> List[date]:
    days, current = [], start
    while current <= end:
        if is_working_day(current):
            days.append(current)
        current += timedelta(days=1)
    return days

def get_current_week_range():
    today = datetime.now(MYT).date()
    monday = today - timedelta(days=today.weekday())
    saturday = monday + timedelta(days=5)
    return monday, saturday

# ─────────────────────────────────────────────
# ATTENDANCE HELPERS
# ─────────────────────────────────────────────

def format_duration_display(minutes: int) -> str:
    if minutes is None or minutes < 0:
        return "0m"
    h = minutes // 60
    m = minutes % 60
    if h > 0:
        return f"{h}h {m}m"
    return f"{m}m"

async def process_auto_clock_outs() -> int:
    """Lazy auto-clock-out: if current MYT time >= auto clock-out time for today,
    find all 'working' attendance records for today and clock them out.
    Auto clock-out time is 18:00 for Mon-Fri, 13:00 for first/last Saturday.
    Returns count of processed records."""
    try:
        now = get_myt_now()
        today_date = now.date()
        cutoff_time = get_auto_clockout_time(today_date)
        if cutoff_time is None:
            return 0  # Not a working day
        if now.time() < cutoff_time:
            return 0  # Not yet time to auto clock-out

        today_str = get_myt_today()
        cutoff_dt = datetime.combine(today_date, cutoff_time, tzinfo=MYT)

        working_records = await db.attendance.find({
            "date": today_str,
            "status": "working"
        }).to_list(1000)

        count = 0
        for rec in working_records:
            try:
                clock_in_at = datetime.fromisoformat(rec["clock_in_at"])
                duration_minutes = int((cutoff_dt - clock_in_at).total_seconds() // 60)
            except (KeyError, ValueError, TypeError):
                duration_minutes = 0
            await db.attendance.update_one(
                {"_id": rec["_id"]},
                {"$set": {
                    "clock_out": cutoff_time.strftime("%H:%M:%S"),
                    "clock_out_at": cutoff_dt.isoformat(),
                    "working_duration_minutes": duration_minutes,
                    "working_duration_display": format_duration_display(duration_minutes),
                    "clock_out_reason": "auto",
                    "status": "completed"
                }}
            )
            count += 1

        if count:
            logger.info(f"Auto clock-out processed {count} records for {today_str} at {cutoff_time}")
        return count
    except Exception as e:
        logger.error(f"process_auto_clock_outs error: {e}")
        return 0

# ─────────────────────────────────────────────
# Dependencies
# ─────────────────────────────────────────────

async def get_current_user(request: Request):
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.profiles.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        result = doc_to_dict(user)
        result.pop("password_hash", None)
        return result
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")

async def require_active(current_user: dict = Depends(get_current_user)):
    if current_user.get("status") != "active":
        raise HTTPException(status_code=403, detail="Account not active")
    return current_user

async def require_admin(current_user: dict = Depends(require_active)):
    if current_user.get("role") not in ["admin", "boss"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

# ─────────────────────────────────────────────
# Startup / Shutdown
# ─────────────────────────────────────────────

@app.on_event("startup")
async def startup():
    await db.profiles.create_index("email", unique=True)
    await db.daily_reports.create_index([("employee_id", 1), ("report_date", 1)], unique=True)
    await db.daily_reports.create_index([("report_date", 1), ("employee_id", 1)])
    await db.daily_reports.create_index([("upload_source", 1), ("report_date", -1)])
    await db.daily_reports.create_index([("employee_id", 1), ("upload_source", 1), ("report_date", -1)])
    await db.login_attempts.create_index("email")
    await db.login_attempts.create_index("attempted_at")
    await db.attendance.create_index([("employee_id", 1), ("date", 1)], unique=True)
    await db.attendance.create_index([("date", 1), ("employee_id", 1)])
    await db.attendance.create_index([("date", 1)])
    await db.leave_requests.create_index([("employee_id", 1), ("date_from", -1)])
    await db.leave_requests.create_index([("status", 1), ("date_from", -1)])
    await db.leave_requests.create_index([("date_from", 1), ("date_to", 1)])
    await db.meetings.create_index([("employee_id", 1), ("start_at", -1)])
    await db.meetings.create_index([("status", 1), ("start_at", 1)])
    await db.meetings.create_index([("start_at", 1)])
    await db.channels.create_index("name", unique=True, sparse=True)
    await db.channels.create_index("members")
    try:
        await db.channels.drop_index("dm_between_1")
    except Exception:
        pass
    # No create_index for dm_between — unique multikey
    # indexes on array fields don't work for this use case
    await db.messages.create_index([("channel_id", 1), ("created_at", 1)])
    await db.messages.create_index("sender_id")
    await db.messages.create_index([("channel_id", 1), ("deleted", 1)])
    await db.message_reads.create_index([("user_id", 1), ("channel_id", 1)], unique=True)

    # One-time migration: remove dm_between from non-DM channels (fixes bad seed data)
    try:
        await db.channels.update_many(
            {"type": {"$in": ["public", "private"]}, "dm_between": None},
            {"$unset": {"dm_between": ""}}
        )
    except Exception as e:
        logger.error(f"Channel migration (unset dm_between) failed: {e}")

    # Seed default channels (idempotent, non-fatal on error)
    now_iso = datetime.now(timezone.utc).isoformat()
    try:
        try:
            general = await db.channels.find_one({"name": "General", "type": "public"})
            if not general:
                await db.channels.insert_one({
                    "name": "General",
                    "type": "public",
                    "members": [],
                    "created_by": "system",
                    "created_at": now_iso
                })
        except Exception as e:
            logger.error(f"Channel seeding failed (General): {e}")

        try:
            management = await db.channels.find_one({"name": "Management", "type": "private"})
            if not management:
                try:
                    admins = await db.profiles.find(
                        {"role": {"$in": ["admin", "boss"]}, "status": "active"},
                        {"_id": 1}
                    ).to_list(100)
                    admin_ids = [str(a["_id"]) for a in admins]
                except Exception as e:
                    logger.error(f"Channel seeding failed (querying admin/boss users for Management): {e}")
                    admin_ids = []
                await db.channels.insert_one({
                    "name": "Management",
                    "type": "private",
                    "members": admin_ids,
                    "created_by": "system",
                    "created_at": now_iso
                })
        except Exception as e:
            logger.error(f"Channel seeding failed (Management): {e}")

        try:
            meeting_alerts = await db.channels.find_one({"name": "Meeting Alerts", "type": "public"})
            if not meeting_alerts:
                await db.channels.insert_one({
                    "name": "Meeting Alerts",
                    "type": "public",
                    "members": [],
                    "created_by": "system",
                    "created_at": now_iso
                })
        except Exception as e:
            logger.error(f"Channel seeding failed (Meeting Alerts): {e}")
    except Exception as e:
        logger.error(f"Channel seeding failed: {e}")

    logger.info("Performance Pulse backend started.")

@app.on_event("shutdown")
async def shutdown():
    client.close()

# ─────────────────────────────────────────────
# BRUTE FORCE HELPERS
# ─────────────────────────────────────────────

async def check_brute_force(email: str):
    """Raise 429 if this email has 5+ failed attempts in the last 15 minutes."""
    cutoff = (datetime.now(timezone.utc) - timedelta(minutes=LOCKOUT_MINUTES)).isoformat()
    count = await db.login_attempts.count_documents({
        "email": email,
        "attempted_at": {"$gt": cutoff}
    })
    if count >= LOCKOUT_ATTEMPTS:
        raise HTTPException(
            status_code=429,
            detail="Too many failed attempts. Please try again after 15 minutes."
        )

async def record_failed_attempt(email: str):
    await db.login_attempts.insert_one({
        "email": email,
        "attempted_at": datetime.now(timezone.utc).isoformat()
    })

async def clear_failed_attempts(email: str):
    await db.login_attempts.delete_many({"email": email})

# ─────────────────────────────────────────────
# AUTH ROUTES
# ─────────────────────────────────────────────

@api_router.post("/auth/bootstrap-boss")
async def bootstrap_boss(request: Request, response: Response):
    """
    One-time endpoint to promote an existing account to boss.
    Requires BOOTSTRAP_SECRET env var to be set.
    Blocked if a boss account already exists (safe to leave enabled).
    """
    bootstrap_secret = os.environ.get("BOOTSTRAP_SECRET", "")
    if not bootstrap_secret:
        raise HTTPException(status_code=403, detail="Bootstrap is not enabled on this server")

    body = await request.json()
    secret = body.get("secret", "")
    email = body.get("email", "").strip().lower()
    full_name = body.get("full_name", "Mr. Seelaan").strip()

    if secret != bootstrap_secret:
        raise HTTPException(status_code=403, detail="Invalid bootstrap secret")
    if not email:
        raise HTTPException(status_code=400, detail="Email is required")

    # Block if boss already exists
    existing_boss = await db.profiles.find_one({"role": "boss"})
    if existing_boss:
        raise HTTPException(status_code=409, detail="A boss account already exists. Bootstrap is only allowed once.")

    user = await db.profiles.find_one({"email": email})
    if not user:
        raise HTTPException(status_code=404, detail="Account not found. Register via the app first, then call this endpoint.")

    await db.profiles.update_one(
        {"email": email},
        {"$set": {"role": "boss", "status": "active", "full_name": full_name}}
    )
    updated = await db.profiles.find_one({"email": email}, {"password_hash": 0})
    return {
        "message": "Boss account created successfully",
        "id": str(updated["_id"]),
        "full_name": updated["full_name"],
        "email": updated["email"],
        "role": updated["role"],
        "status": updated["status"]
    }

@api_router.post("/auth/register")
async def register(request: Request, response: Response):
    body = await request.json()
    email = body.get("email", "").strip().lower()
    password = body.get("password", "")
    full_name = body.get("full_name", "").strip()
    department = body.get("department", "")
    job_title = body.get("job_title", "").strip()

    if not email or not password or not full_name:
        raise HTTPException(status_code=400, detail="Email, password, and full name are required")
    if len(password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters")

    existing = await db.profiles.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=409, detail="Email already registered")

    user_doc = {
        "full_name": full_name,
        "email": email,
        "password_hash": hash_password(password),
        "role": "employee",
        "status": "pending",
        "department": department,
        "job_title": job_title,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.profiles.insert_one(user_doc)
    user_id = str(result.inserted_id)

    access_token = create_access_token(user_id, email)
    refresh_token = create_refresh_token(user_id)
    set_auth_cookies(response, access_token, refresh_token)

    return {"id": user_id, "full_name": full_name, "email": email,
            "role": "employee", "status": "pending", "department": department, "job_title": job_title}

@api_router.post("/auth/login")
async def login(request: Request, response: Response):
    body = await request.json()
    email = body.get("email", "").strip().lower()
    password = body.get("password", "")

    if not email or not password:
        raise HTTPException(status_code=400, detail="Email and password are required")

    # Brute force check — before any DB user lookup
    await check_brute_force(email)

    user = await db.profiles.find_one({"email": email})
    if not user or not verify_password(password, user.get("password_hash", "")):
        await record_failed_attempt(email)
        raise HTTPException(status_code=401, detail="Invalid email or password")

    # Successful login — clear any previous failed attempts
    await clear_failed_attempts(email)

    user_dict = doc_to_dict(user)
    user_dict.pop("password_hash", None)
    user_id = user_dict["id"]

    access_token = create_access_token(user_id, email)
    refresh_token = create_refresh_token(user_id)
    set_auth_cookies(response, access_token, refresh_token)

    return user_dict

@api_router.post("/auth/logout")
async def logout(response: Response, current_user: dict = Depends(get_current_user)):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Logged out"}

@api_router.get("/auth/me")
async def me(current_user: dict = Depends(get_current_user)):
    return current_user

@api_router.post("/auth/refresh")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="No refresh token")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.profiles.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user_id = str(user["_id"])
        email = user.get("email", "")
        new_token = create_access_token(user_id, email)
        response.set_cookie("access_token", new_token, httponly=True, secure=COOKIE_SECURE, samesite=COOKIE_SAMESITE, max_age=28800, path="/")
        return {"message": "Token refreshed"}
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid refresh token")

# ─────────────────────────────────────────────
# USER ROUTES
# ─────────────────────────────────────────────

@api_router.get("/users")
async def list_users(
    status: Optional[str] = None,
    role: Optional[str] = None,
    current_user: dict = Depends(require_admin)
):
    query = {}
    if status:
        query["status"] = status
    if role:
        query["role"] = role
    users = await db.profiles.find(query, {"password_hash": 0}).sort("created_at", -1).to_list(1000)
    return [doc_to_dict(u) for u in users]

@api_router.get("/users/departments")
async def get_departments(current_user: dict = Depends(get_current_user)):
    return {"departments": DEPARTMENTS}

@api_router.get("/users/dm-list")
async def get_dm_list(current_user: dict = Depends(require_active)):
    users = await db.profiles.find(
        {"status": "active", "_id": {"$ne": ObjectId(current_user["id"])}},
        {"password_hash": 0}
    ).sort("full_name", 1).to_list(1000)
    return [
        {
            "id": str(u["_id"]),
            "full_name": u.get("full_name", ""),
            "role": u.get("role", ""),
            "department": u.get("department", ""),
        }
        for u in users
    ]

@api_router.get("/users/{user_id}")
async def get_user(user_id: str, current_user: dict = Depends(require_active)):
    if current_user["id"] != user_id and current_user["role"] not in ["admin", "boss"]:
        raise HTTPException(status_code=403, detail="Access denied")
    try:
        user = await db.profiles.find_one({"_id": ObjectId(user_id)}, {"password_hash": 0})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid user ID")
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return doc_to_dict(user)

@api_router.put("/users/{user_id}/status")
async def update_user_status(user_id: str, request: Request, current_user: dict = Depends(require_admin)):
    body = await request.json()
    new_status = body.get("status")
    if new_status not in ["active", "inactive", "rejected", "pending"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    try:
        result = await db.profiles.update_one({"_id": ObjectId(user_id)}, {"$set": {"status": new_status}})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid user ID")
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")

    if new_status == "active":
        try:
            general = await db.channels.find_one(
                {"name": "General", "type": "public"}
            )
            if general:
                await db.channels.update_one(
                    {"_id": general["_id"]},
                    {"$addToSet": {"members": user_id}}
                )
        except Exception as e:
            logger.error(f"Auto-add to General channel failed: {e}")

    return {"message": f"Status updated to {new_status}"}

@api_router.put("/users/{user_id}/role")
async def update_user_role(user_id: str, request: Request, current_user: dict = Depends(require_admin)):
    body = await request.json()
    new_role = body.get("role")
    if new_role not in ["employee", "admin", "boss"]:
        raise HTTPException(status_code=400, detail="Invalid role")
    try:
        result = await db.profiles.update_one({"_id": ObjectId(user_id)}, {"$set": {"role": new_role}})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid user ID")
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")

    return {"message": f"Role updated to {new_role}"}

@api_router.put("/users/{user_id}")
async def update_user_profile(user_id: str, request: Request, current_user: dict = Depends(require_active)):
    if current_user["id"] != user_id and current_user["role"] not in ["admin", "boss"]:
        raise HTTPException(status_code=403, detail="Access denied")
    body = await request.json()
    allowed = ["full_name", "department", "job_title", "phone", "profile_remarks"]
    update_data = {k: v for k, v in body.items() if k in allowed}
    if not update_data:
        raise HTTPException(status_code=400, detail="No valid fields to update")
    try:
        result = await db.profiles.update_one({"_id": ObjectId(user_id)}, {"$set": update_data})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid user ID")
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")

    return {"message": "Profile updated"}

# ─────────────────────────────────────────────
# REPORT ROUTES
# ─────────────────────────────────────────────

@api_router.get("/reports/template")
async def download_template(current_user: dict = Depends(require_active)):
    xlsx = generate_excel_template(current_user.get("full_name", ""), current_user.get("department", ""))
    return Response(
        content=xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=daily_report_{get_myt_today()}.xlsx"}
    )

@api_router.post("/reports/upload-preview")
async def upload_report_preview(file: UploadFile = File(...), current_user: dict = Depends(require_active)):
    if current_user["role"] != "employee":
        raise HTTPException(status_code=403, detail="Only employees can upload reports")
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

    content = await file.read()
    try:
        rows = parse_daily_report_excel(content)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    if not rows:
        raise HTTPException(status_code=400, detail="No data found. Ensure row 5 onwards contains report data.")

    today_str = get_myt_today()
    today_date = date.fromisoformat(today_str)

    if not is_working_day(today_date):
        raise HTTPException(status_code=400, detail="Reports can only be submitted on working days (Monday to Saturday).")

    for row in rows:
        if row["report_date"] != today_str:
            raise HTTPException(
                status_code=400,
                detail=f"Report date '{row['report_date']}' does not match today ({today_str} MYT). Upload rejected."
            )

    existing = await db.daily_reports.find_one({"employee_id": current_user["id"], "report_date": today_str})
    if existing:
        raise HTTPException(status_code=409, detail="You have already submitted a report for today.")

    return {"preview": rows, "today": today_str, "filename": file.filename, "row_count": len(rows)}

@api_router.post("/reports/upload-confirm")
async def upload_report_confirm(file: UploadFile = File(...), current_user: dict = Depends(require_active)):
    if current_user["role"] != "employee":
        raise HTTPException(status_code=403, detail="Only employees can upload reports")
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

    content = await file.read()
    try:
        rows = parse_daily_report_excel(content)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    if not rows:
        raise HTTPException(status_code=400, detail="No data found in the Excel file.")

    today_str = get_myt_today()
    today_date = date.fromisoformat(today_str)

    if not is_working_day(today_date):
        raise HTTPException(status_code=400, detail="Reports can only be submitted on working days.")

    for row in rows:
        if row["report_date"] != today_str:
            raise HTTPException(status_code=400, detail=f"Report date mismatch. Expected {today_str}.")

    existing = await db.daily_reports.find_one({"employee_id": current_user["id"], "report_date": today_str})
    if existing:
        raise HTTPException(status_code=409, detail="You have already submitted a report for today.")

    # Upload file to object storage
    file_path = f"{APP_STORAGE_PREFIX}/reports/{current_user['id']}/{today_str}_{uuid.uuid4().hex[:8]}.xlsx"
    stored_path = None
    try:
        result = put_object(file_path, content,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        stored_path = result.get("path", file_path)
    except Exception as exc:
        logger.warning(f"Storage upload failed (proceeding without file): {exc}")

    # Aggregate multiple rows into one report
    primary = rows[0]
    if len(rows) > 1:
        morning_plan = "\n".join(r["morning_plan"] for r in rows if r["morning_plan"])
        afternoon_plan = "\n".join(r["afternoon_plan"] for r in rows if r["afternoon_plan"])
        final_report = "\n".join(r["final_report"] for r in rows if r["final_report"])
        statuses = [r["task_status"] for r in rows]
        for worst in ["Delayed", "Pending", "In Progress", "Completed"]:
            if worst in statuses:
                combined_status = worst
                break
        else:
            combined_status = "Completed"
    else:
        morning_plan = primary["morning_plan"]
        afternoon_plan = primary["afternoon_plan"]
        final_report = primary["final_report"]
        combined_status = primary["task_status"]

    report_doc = {
        "employee_id": current_user["id"],
        "employee_name": current_user.get("full_name", ""),
        "report_date": today_str,
        "morning_plan": morning_plan,
        "afternoon_plan": afternoon_plan,
        "final_report": final_report,
        "task_category": primary["task_category"],
        "task_status": combined_status,
        "calls_made": sum(r.get("calls_made", 0) for r in rows),
        "follow_ups": sum(r.get("follow_ups", 0) for r in rows),
        "interested_leads": sum(r.get("interested_leads", 0) for r in rows),
        "blockers": primary.get("blockers", ""),
        "final_remarks": primary.get("final_remarks", ""),
        "morning_plan_matched": primary.get("morning_plan_matched", ""),
        "afternoon_plan_matched": primary.get("afternoon_plan_matched", ""),
        "overall_tally_status": primary.get("overall_tally_status", ""),
        "review_status": "submitted",
        "upload_source": "excel",
        "original_filename": file.filename,
        "file_path": stored_path,
        "submitted_after_6pm": get_myt_now().hour >= 20,
        "raw_rows": rows,
        "created_at": datetime.now(timezone.utc).isoformat()
    }

    ins = await db.daily_reports.insert_one(report_doc)
    report_doc["id"] = str(ins.inserted_id)
    report_doc.pop("_id", None)
    return report_doc

@api_router.get("/files/{file_path:path}")
async def serve_file(file_path: str, current_user: dict = Depends(require_active)):
    try:
        data, content_type = get_object(file_path)
        basename = file_path.split("/")[-1]
        filename = basename.split("_", 1)[-1] if "_" in basename else basename
        disposition = "inline" if content_type.startswith("image/") else "attachment"
        return Response(content=data, media_type=content_type,
                        headers={"Content-Disposition": f'{disposition}; filename="{filename}"'})
    except Exception:
        raise HTTPException(status_code=404, detail="File not found")

@api_router.get("/reports/my")
async def get_my_reports(current_user: dict = Depends(require_active)):
    reports = await db.daily_reports.find(
        {"employee_id": current_user["id"]}
    ).sort("report_date", -1).to_list(200)
    return [doc_to_dict(r) for r in reports]

@api_router.get("/reports/today")
async def get_today_report(current_user: dict = Depends(require_active)):
    today_str = get_myt_today()
    report = await db.daily_reports.find_one({"employee_id": current_user["id"], "report_date": today_str})
    return {"report": doc_to_dict(report), "today": today_str, "is_working_day": is_working_day(date.fromisoformat(today_str))}

@api_router.get("/reports/employee/{employee_id}")
async def get_employee_reports(employee_id: str, current_user: dict = Depends(require_admin)):
    reports = await db.daily_reports.find(
        {"employee_id": employee_id}
    ).sort("report_date", -1).to_list(200)
    return [doc_to_dict(r) for r in reports]

@api_router.post("/reports")
async def submit_report(request: Request, current_user: dict = Depends(require_active)):
    if current_user["role"] != "employee":
        raise HTTPException(status_code=403, detail="Only employees can submit reports")

    body = await request.json()
    today_str = get_myt_today()
    today = date.fromisoformat(today_str)

    if not is_working_day(today):
        raise HTTPException(status_code=400, detail="Reports can only be submitted on working days (Monday to Saturday)")

    existing = await db.daily_reports.find_one({"employee_id": current_user["id"], "report_date": today_str})
    if existing:
        raise HTTPException(status_code=409, detail="You have already submitted a report for today")

    if not body.get("morning_plan") or not body.get("final_report"):
        raise HTTPException(status_code=400, detail="Morning plan and final report are required")

    task_category = body.get("task_category", "Other")
    task_status_val = body.get("task_status", "Completed")
    if task_category not in TASK_CATEGORIES:
        task_category = "Other"
    if task_status_val not in TASK_STATUSES:
        task_status_val = "Completed"

    report_doc = {
        "employee_id": current_user["id"],
        "employee_name": current_user.get("full_name", ""),
        "report_date": today_str,
        "morning_plan": body.get("morning_plan", ""),
        "afternoon_plan": body.get("afternoon_plan", ""),
        "final_report": body.get("final_report", ""),
        "task_category": task_category,
        "task_status": task_status_val,
        "calls_made": max(0, int(body.get("calls_made", 0))),
        "follow_ups": max(0, int(body.get("follow_ups", 0))),
        "interested_leads": max(0, int(body.get("interested_leads", 0))),
        "blockers": body.get("blockers", ""),
        "final_remarks": body.get("final_remarks", ""),
        "review_status": "submitted",
        "submitted_after_6pm": get_myt_now().hour >= 20,
        "created_at": datetime.now(timezone.utc).isoformat()
    }

    result = await db.daily_reports.insert_one(report_doc)
    report_doc["id"] = str(result.inserted_id)
    report_doc.pop("_id", None)

    return report_doc

@api_router.get("/reports")
async def get_all_reports(
    employee_id: Optional[str] = None,
    report_date: Optional[str] = None,
    task_category: Optional[str] = None,
    task_status: Optional[str] = None,
    review_status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(require_admin)
):
    query = {}
    if employee_id:
        query["employee_id"] = employee_id
    if report_date:
        query["report_date"] = report_date
    if task_category:
        query["task_category"] = task_category
    if task_status:
        query["task_status"] = task_status
    if review_status:
        query["review_status"] = review_status
    if date_from or date_to:
        date_q = {}
        if date_from:
            date_q["$gte"] = date_from
        if date_to:
            date_q["$lte"] = date_to
        query["report_date"] = date_q

    reports = await db.daily_reports.find(query).sort("report_date", -1).to_list(1000)
    return [doc_to_dict(r) for r in reports]

@api_router.get("/reports/{report_id}")
async def get_report(report_id: str, current_user: dict = Depends(require_active)):
    try:
        report = await db.daily_reports.find_one({"_id": ObjectId(report_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid report ID")
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    if current_user["role"] == "employee" and report.get("employee_id") != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    return doc_to_dict(report)

@api_router.get("/reports/{report_id}/download")
async def download_report_excel(report_id: str, current_user: dict = Depends(require_active)):
    try:
        report = await db.daily_reports.find_one({"_id": ObjectId(report_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid report ID")
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    if current_user["role"] == "employee" and report.get("employee_id") != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    employee_name = report.get("employee_name", "Employee")
    report_date = report.get("report_date", "unknown")
    filename = f"Report_{employee_name}_{report_date}.xlsx"

    if report.get("upload_source") == "excel" and report.get("file_path"):
        try:
            data, content_type = get_object(report["file_path"])
            return Response(
                content=data,
                media_type=content_type,
                headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
            )
        except Exception:
            pass

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"

    headers = ["Field", "Value"]
    ws.append(headers)

    data_rows = [
        ["Employee", employee_name],
        ["Report Date", report_date],
        ["Task Category", report.get("task_category", "")],
        ["Task Status", report.get("task_status", "")],
        ["Morning Plan", report.get("morning_plan", "")],
        ["Afternoon Plan", report.get("afternoon_plan", "")],
        ["Final Report", report.get("final_report", "")],
        ["Calls Made", report.get("calls_made", 0)],
        ["Follow-ups", report.get("follow_ups", 0)],
        ["Interested Leads", report.get("interested_leads", 0)],
        ["Blockers", report.get("blockers", "")],
        ["Final Remarks", report.get("final_remarks", "")],
        ["Submitted At", report.get("created_at", "")],
    ]
    for row in data_rows:
        ws.append(row)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

@api_router.get("/reports/history")
async def get_reports_history(
    start_date: str,
    end_date: str,
    employee_id: Optional[str] = None,
    status: Optional[str] = "all",
    current_user: dict = Depends(require_admin)
):
    try:
        from bson import ObjectId as _OID

        active_employees = await db.profiles.find(
            {"role": "employee", "status": "active"}, {"password_hash": 0}
        ).to_list(1000)

        if employee_id:
            active_employees = [e for e in active_employees if str(e["_id"]) == employee_id]

        emp_map = {str(e["_id"]): e for e in active_employees}

        query = {"report_date": {"$gte": start_date, "$lte": end_date}}
        if employee_id:
            query["employee_id"] = employee_id

        reports = await db.daily_reports.find(query).to_list(2000)
        report_map = {}
        for r in reports:
            key = f"{r.get('employee_id')}_{r.get('report_date')}"
            report_map[key] = r

        start_d = date.fromisoformat(start_date)
        end_d = date.fromisoformat(end_date)

        result = []
        cur = start_d
        while cur <= end_d:
            if not is_working_day(cur):
                cur += timedelta(days=1)
                continue
            cur_str = cur.isoformat()
            for emp in active_employees:
                eid = str(emp["_id"])
                key = f"{eid}_{cur_str}"
                rep = report_map.get(key)
                if rep:
                    if status == "missing":
                        continue
                    result.append({
                        "id": str(rep.get("_id", "")),
                        "employee_id": eid,
                        "employee_name": emp.get("full_name", rep.get("employee_name", "")),
                        "department": emp.get("department", ""),
                        "report_date": cur_str,
                        "task_category": rep.get("task_category", ""),
                        "task_status": rep.get("task_status", ""),
                        "calls_made": rep.get("calls_made", 0),
                        "follow_ups": rep.get("follow_ups", 0),
                        "interested_leads": rep.get("interested_leads", 0),
                        "review_status": rep.get("review_status", ""),
                        "submitted_at": rep.get("created_at", ""),
                        "submitted_after_6pm": rep.get("submitted_after_6pm", False),
                        "upload_source": rep.get("upload_source", ""),
                        "original_filename": rep.get("original_filename", ""),
                        "file_path": rep.get("file_path", ""),
                        "report_status": "submitted"
                    })
                else:
                    if status == "submitted":
                        continue
                    result.append({
                        "id": None,
                        "employee_id": eid,
                        "employee_name": emp.get("full_name", ""),
                        "department": emp.get("department", ""),
                        "report_date": cur_str,
                        "task_category": None,
                        "task_status": None,
                        "calls_made": None,
                        "follow_ups": None,
                        "interested_leads": None,
                        "review_status": None,
                        "submitted_at": None,
                        "submitted_after_6pm": None,
                        "upload_source": None,
                        "original_filename": None,
                        "report_status": "missing"
                    })
            cur += timedelta(days=1)

        result.sort(key=lambda x: (x["report_date"], x["employee_name"]), reverse=True)
        return result
    except Exception as e:
        logger.error(f"Error in get_reports_history: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/reports/history/summary")
async def get_reports_history_summary(
    start_date: str,
    end_date: str,
    employee_id: Optional[str] = None,
    current_user: dict = Depends(require_admin)
):
    try:
        active_employees = await db.profiles.find(
            {"role": "employee", "status": "active"}, {"_id": 1, "full_name": 1}
        ).to_list(1000)

        if employee_id:
            active_employees = [e for e in active_employees if str(e["_id"]) == employee_id]

        emp_count = len(active_employees)

        start_d = date.fromisoformat(start_date)
        end_d = date.fromisoformat(end_date)
        working_days = 0
        cur = start_d
        while cur <= end_d:
            if is_working_day(cur):
                working_days += 1
            cur += timedelta(days=1)

        total_expected = working_days * emp_count

        query = {"report_date": {"$gte": start_date, "$lte": end_date}}
        if employee_id:
            query["employee_id"] = employee_id

        reports = await db.daily_reports.find(query).to_list(2000)
        total_submitted = len(reports)
        total_missing = max(0, total_expected - total_submitted)
        submission_rate = round(total_submitted / total_expected * 100, 1) if total_expected > 0 else 0

        emp_report_counts = {}
        for r in reports:
            eid = r.get("employee_id")
            if eid:
                emp_report_counts[eid] = emp_report_counts.get(eid, 0) + 1

        top_performer = None
        if emp_report_counts and working_days > 0:
            top_eid = max(emp_report_counts, key=emp_report_counts.get)
            top_count = emp_report_counts[top_eid]
            top_emp = next((e for e in active_employees if str(e["_id"]) == top_eid), None)
            top_name = top_emp.get("full_name", "Unknown") if top_emp else "Unknown"
            top_rate = round(top_count / working_days * 100, 1)
            top_performer = {"name": top_name, "submission_count": top_count, "rate": top_rate}

        return {
            "total_expected": total_expected,
            "total_submitted": total_submitted,
            "total_missing": total_missing,
            "submission_rate": submission_rate,
            "top_performer": top_performer
        }
    except Exception as e:
        logger.error(f"Error in get_reports_history_summary: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/reports/{report_id}/review")
async def review_report(report_id: str, request: Request, current_user: dict = Depends(require_admin)):
    body = await request.json()
    review_status = body.get("review_status")
    if review_status not in REVIEW_STATUSES:
        raise HTTPException(status_code=400, detail="Invalid review status")
    try:
        result = await db.daily_reports.update_one(
            {"_id": ObjectId(report_id)},
            {"$set": {"review_status": review_status, "reviewed_by": current_user["id"],
                      "reviewed_at": datetime.now(timezone.utc).isoformat()}}
        )
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid report ID")
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Report not found")
    return {"message": f"Report marked as {review_status}"}

# ─────────────────────────────────────────────
# DASHBOARD ROUTES
# ─────────────────────────────────────────────

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: dict = Depends(require_admin)):
    await process_auto_clock_outs()
    today_str = get_myt_today()
    today_date = date.fromisoformat(today_str)
    monday, saturday = get_current_week_range()
    week_start_str = monday.isoformat()
    week_end_str = saturday.isoformat()

    total_employees = await db.profiles.count_documents({"role": "employee"})
    active_employees = await db.profiles.count_documents({"role": "employee", "status": "active"})
    pending_approvals = await db.profiles.count_documents({"status": "pending"})
    reports_today = await db.daily_reports.count_documents({"report_date": today_str})

    missing_count = 0
    if is_working_day(today_date):
        all_active = await db.profiles.find(
            {"role": "employee", "status": "active"}, {"_id": 1}
        ).to_list(1000)
        reported_today = await db.daily_reports.find(
            {"report_date": today_str}, {"employee_id": 1}
        ).to_list(1000)
        reported_ids = {r["employee_id"] for r in reported_today}
        missing_count = sum(1 for e in all_active if str(e["_id"]) not in reported_ids)

    weekly_reports = await db.daily_reports.find({
        "report_date": {"$gte": week_start_str, "$lte": week_end_str}
    }).to_list(1000)

    return {
        "total_employees": total_employees,
        "active_employees": active_employees,
        "pending_approvals": pending_approvals,
        "reports_today": reports_today,
        "missing_today": missing_count,
        "total_calls_this_week": sum(r.get("calls_made", 0) for r in weekly_reports),
        "total_followups_this_week": sum(r.get("follow_ups", 0) for r in weekly_reports),
        "total_leads_this_week": sum(r.get("interested_leads", 0) for r in weekly_reports),
        "today": today_str,
        "is_working_day": is_working_day(today_date),
        "week_start": week_start_str,
        "week_end": week_end_str
    }

@api_router.get("/dashboard/boss-charts")
async def get_boss_charts(current_user: dict = Depends(require_admin)):
    """Returns aggregated chart data for Boss Dashboard visualizations."""
    await process_auto_clock_outs()
    monday, saturday = get_current_week_range()
    week_start_str = monday.isoformat()
    week_end_str = saturday.isoformat()

    # Task status breakdown for current week
    weekly_reports = await db.daily_reports.find({
        "report_date": {"$gte": week_start_str, "$lte": week_end_str}
    }).to_list(1000)

    task_breakdown = [
        {"status": "Completed", "count": sum(1 for r in weekly_reports if r.get("task_status") == "Completed"), "fill": "#10b981"},
        {"status": "In Progress", "count": sum(1 for r in weekly_reports if r.get("task_status") == "In Progress"), "fill": "#3b82f6"},
        {"status": "Pending", "count": sum(1 for r in weekly_reports if r.get("task_status") == "Pending"), "fill": "#f59e0b"},
        {"status": "Delayed", "count": sum(1 for r in weekly_reports if r.get("task_status") == "Delayed"), "fill": "#ef4444"},
    ]

    # Weekly score trend — last 6 weeks
    score_trend = []
    current_monday = monday
    for _ in range(6):
        w_start = current_monday
        w_end = current_monday + timedelta(days=5)
        summaries = await db.weekly_summaries.find({
            "week_start": w_start.isoformat(),
            "week_end": w_end.isoformat()
        }).to_list(200)
        avg_score = round(sum(s.get("performance_score", 0) for s in summaries) / len(summaries), 1) if summaries else 0
        score_trend.append({
            "week": w_start.strftime("%-d %b"),
            "avg_score": avg_score,
            "count": len(summaries)
        })
        current_monday = current_monday - timedelta(days=7)
    score_trend.reverse()

    return {
        "task_breakdown": task_breakdown,
        "score_trend": score_trend,
        "week_start": week_start_str,
        "week_end": week_end_str,
    }

@api_router.get("/dashboard/missing-today")
async def get_missing_today(current_user: dict = Depends(require_admin)):
    await process_auto_clock_outs()
    today_str = get_myt_today()
    today_date = date.fromisoformat(today_str)

    if not is_working_day(today_date):
        return {"is_working_day": False, "today": today_str, "missing_employees": []}

    active_employees = await db.profiles.find(
        {"role": "employee", "status": "active"}, {"password_hash": 0}
    ).to_list(1000)
    reported_today = await db.daily_reports.find(
        {"report_date": today_str}, {"employee_id": 1}
    ).to_list(1000)
    reported_ids = {r["employee_id"] for r in reported_today}
    missing = [doc_to_dict(e) for e in active_employees if str(e["_id"]) not in reported_ids]

    return {"is_working_day": True, "today": today_str, "missing_employees": missing}

# ─────────────────────────────────────────────
# WEEKLY SUMMARY ROUTES
# ─────────────────────────────────────────────

async def _generate_weekly_summary(employee_id: str, week_start_str: str, week_end_str: str) -> str:
    reports = await db.daily_reports.find({
        "employee_id": employee_id,
        "report_date": {"$gte": week_start_str, "$lte": week_end_str}
    }).to_list(100)

    start = date.fromisoformat(week_start_str)
    end = date.fromisoformat(week_end_str)
    today = datetime.now(MYT).date()
    past_working_days = [d for d in get_working_days_in_range(start, end) if d <= today]

    total_submitted = len(reports)
    missing_days = max(0, len(past_working_days) - total_submitted)

    completed = sum(1 for r in reports if r.get("task_status") == "Completed")
    in_progress = sum(1 for r in reports if r.get("task_status") == "In Progress")
    pending = sum(1 for r in reports if r.get("task_status") == "Pending")
    delayed = sum(1 for r in reports if r.get("task_status") == "Delayed")

    total_calls = sum(r.get("calls_made", 0) for r in reports)
    total_follow_ups = sum(r.get("follow_ups", 0) for r in reports)
    total_leads = sum(r.get("interested_leads", 0) for r in reports)

    parts = [
        f"[{r['report_date']}] {r.get('final_report', '').strip()}"
        for r in sorted(reports, key=lambda x: x["report_date"])
        if r.get("final_report", "").strip()
    ]
    compiled = " | ".join(parts) if parts else "No reports submitted this week."

    summary_doc = {
        "employee_id": employee_id,
        "week_start": week_start_str,
        "week_end": week_end_str,
        "total_reports_submitted": total_submitted,
        "working_days_in_period": len(past_working_days),
        "missing_days": missing_days,
        "completed_count": completed,
        "in_progress_count": in_progress,
        "pending_count": pending,
        "delayed_count": delayed,
        "total_calls": total_calls,
        "total_follow_ups": total_follow_ups,
        "total_interested_leads": total_leads,
        "compiled_summary": compiled,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        **calculate_performance_score(total_submitted, len(past_working_days), completed,
                                       delayed, missing_days, total_calls)
    }

    existing = await db.weekly_summaries.find_one({
        "employee_id": employee_id,
        "week_start": week_start_str,
        "week_end": week_end_str
    })

    if existing:
        await db.weekly_summaries.update_one({"_id": existing["_id"]}, {"$set": summary_doc})
        return str(existing["_id"])
    else:
        result = await db.weekly_summaries.insert_one(summary_doc)
        return str(result.inserted_id)

@api_router.post("/weekly-summaries/generate")
async def generate_weekly_summary(request: Request, current_user: dict = Depends(require_admin)):
    body = await request.json()
    employee_id = body.get("employee_id")
    week_start_str = body.get("week_start")
    week_end_str = body.get("week_end")

    if not week_start_str or not week_end_str:
        monday, saturday = get_current_week_range()
        week_start_str = monday.isoformat()
        week_end_str = saturday.isoformat()

    if employee_id:
        summary_id = await _generate_weekly_summary(employee_id, week_start_str, week_end_str)
        return {"message": "Summary generated", "id": summary_id, "week_start": week_start_str, "week_end": week_end_str}
    else:
        employees = await db.profiles.find(
            {"role": "employee", "status": "active"}, {"_id": 1}
        ).to_list(1000)
        count = 0
        for emp in employees:
            await _generate_weekly_summary(str(emp["_id"]), week_start_str, week_end_str)
            count += 1
        return {"message": f"Generated summaries for {count} employees", "count": count,
                "week_start": week_start_str, "week_end": week_end_str}

@api_router.post("/weekly-summaries/auto-generate")
async def auto_generate_weekly_summaries(current_user: dict = Depends(require_admin)):
    """Auto-generate summaries if it's Saturday >= 6 PM MYT"""
    now = get_myt_now()
    if now.weekday() != 5 or now.hour < 18:
        return {"auto_generated": False, "reason": "Only runs on Saturday at or after 6 PM MYT",
                "current_day": now.strftime("%A"), "current_hour": now.hour}

    monday, saturday = get_current_week_range()
    week_start_str = monday.isoformat()
    week_end_str = saturday.isoformat()

    employees = await db.profiles.find(
        {"role": "employee", "status": "active"}, {"_id": 1}
    ).to_list(1000)
    count = 0
    for emp in employees:
        await _generate_weekly_summary(str(emp["_id"]), week_start_str, week_end_str)
        count += 1

    return {"auto_generated": True, "count": count, "week_start": week_start_str, "week_end": week_end_str}

@api_router.get("/weekly-summaries")
async def list_weekly_summaries(
    employee_id: Optional[str] = None,
    week_start: Optional[str] = None,
    current_user: dict = Depends(require_admin)
):
    query = {}
    if employee_id:
        query["employee_id"] = employee_id
    if week_start:
        query["week_start"] = week_start
    summaries = await db.weekly_summaries.find(query).sort("week_start", -1).to_list(200)
    return [doc_to_dict(s) for s in summaries]

@api_router.get("/weekly-summaries/employee/{employee_id}")
async def get_employee_summaries(employee_id: str, current_user: dict = Depends(require_admin)):
    summaries = await db.weekly_summaries.find(
        {"employee_id": employee_id}
    ).sort("week_start", -1).to_list(50)
    return [doc_to_dict(s) for s in summaries]

# ─────────────────────────────────────────────
# ATTENDANCE ROUTES
# ─────────────────────────────────────────────

@api_router.post("/attendance/clock-in")
async def attendance_clock_in(current_user: dict = Depends(require_active)):
    if current_user["role"] != "employee":
        raise HTTPException(status_code=403, detail="Only employees can clock in")

    today_str = get_myt_today()
    today_date = date.fromisoformat(today_str)

    if not is_working_day(today_date):
        raise HTTPException(status_code=400, detail="Attendance is not required on non-working days")

    now = get_myt_now()
    if now.hour >= 18:
        raise HTTPException(status_code=400, detail="Clock-in is only allowed before 6:00 PM MYT")

    existing = await db.attendance.find_one({
        "employee_id": current_user["id"],
        "date": today_str
    })
    if existing:
        if existing.get("status") == "completed":
            raise HTTPException(status_code=409, detail="Already completed attendance for today. Cannot clock in again.")
        raise HTTPException(status_code=409, detail="Already clocked in today")

    att_doc = {
        "employee_id": current_user["id"],
        "employee_name": current_user.get("full_name", ""),
        "date": today_str,
        "clock_in": now.strftime("%H:%M:%S"),
        "clock_out": None,
        "clock_in_at": now.isoformat(),
        "clock_out_at": None,
        "working_duration_minutes": 0,
        "working_duration_display": "0m",
        "clock_out_reason": None,
        "status": "working",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.attendance.insert_one(att_doc)
    att_doc["_id"] = result.inserted_id

    return doc_to_dict(att_doc)

@api_router.post("/attendance/clock-out")
async def attendance_clock_out(current_user: dict = Depends(require_active)):
    if current_user["role"] != "employee":
        raise HTTPException(status_code=403, detail="Only employees can clock out")

    today_str = get_myt_today()
    record = await db.attendance.find_one({
        "employee_id": current_user["id"],
        "date": today_str
    })
    if not record:
        raise HTTPException(status_code=404, detail="No clock-in record found for today")
    if record.get("status") == "completed":
        raise HTTPException(status_code=409, detail="You have already clocked out today")

    now = get_myt_now()
    clock_in_at = datetime.fromisoformat(record["clock_in_at"])
    duration_minutes = int((now - clock_in_at).total_seconds() // 60)

    await db.attendance.update_one(
        {"_id": record["_id"]},
        {"$set": {
            "clock_out": now.strftime("%H:%M:%S"),
            "clock_out_at": now.isoformat(),
            "working_duration_minutes": duration_minutes,
            "working_duration_display": format_duration_display(duration_minutes),
            "clock_out_reason": "manual",
            "status": "completed"
        }}
    )

    updated = await db.attendance.find_one({"_id": record["_id"]})
    return doc_to_dict(updated)

@api_router.get("/attendance/today")
async def get_attendance_today(current_user: dict = Depends(require_active)):
    await process_auto_clock_outs()
    today_str = get_myt_today()
    today_date = date.fromisoformat(today_str)
    record = await db.attendance.find_one({
        "employee_id": current_user["id"],
        "date": today_str
    })
    return {
        "attendance": doc_to_dict(record),
        "today": today_str,
        "is_working_day": is_working_day(today_date)
    }

@api_router.get("/attendance/all-today")
async def get_all_attendance_today(
    date: Optional[str] = None,
    current_user: dict = Depends(require_admin)
):
    try:
        try:
            await process_auto_clock_outs()
        except Exception as e:
            logger.warning(f"Auto clock-out error (non-critical): {e}")

        target_date = date if date else get_myt_today()
        try:
            target_date_obj = date.fromisoformat(target_date)
        except Exception:
            target_date_obj = datetime.now(MYT).date()
        working_day = is_working_day(target_date_obj)

        attendance_records = await db.attendance.find(
            {"date": target_date}
        ).to_list(1000)
        att_by_emp = {}
        for r in attendance_records:
            emp_id = r.get("employee_id")
            if emp_id:
                att_by_emp[str(emp_id)] = doc_to_dict(r)

        active_employees = await db.profiles.find(
            {"role": "employee", "status": "active"}, {"password_hash": 0}
        ).to_list(1000)

        attendance_list = []
        not_clocked_in = []
        present_count = 0
        still_working_count = 0
        auto_clocked_out_count = 0

        for emp in active_employees:
            emp_id = str(emp["_id"])
            att = att_by_emp.get(emp_id)
            if att:
                present_count += 1
                if att.get("status") == "working":
                    still_working_count += 1
                if att.get("clock_out_reason") == "auto":
                    auto_clocked_out_count += 1
                attendance_list.append({
                    "employee_id": emp_id,
                    "employee_name": emp.get("full_name", ""),
                    "department": emp.get("department", ""),
                    "clock_in": att.get("clock_in"),
                    "clock_out": att.get("clock_out"),
                    "working_duration_display": att.get("working_duration_display", ""),
                    "clock_out_reason": att.get("clock_out_reason"),
                    "status": att.get("status")
                })
            else:
                not_clocked_in.append({
                    "employee_id": emp_id,
                    "employee_name": emp.get("full_name", ""),
                    "department": emp.get("department", "")
                })

        absent_count = 0 if not working_day else len(not_clocked_in)

        return {
            "date": target_date,
            "is_working_day": working_day,
            "attendance": attendance_list,
            "not_clocked_in": not_clocked_in,
            "summary": {
                "total_employees": len(active_employees),
                "present_today": present_count,
                "absent_today": absent_count,
                "still_working": still_working_count,
                "auto_clocked_out": auto_clocked_out_count
            }
        }
    except Exception as e:
        logger.error(f"Error in get_all_attendance_today: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@api_router.get("/dashboard/attendance-summary")
async def get_attendance_summary(current_user: dict = Depends(require_admin)):
    await process_auto_clock_outs()
    today_str = get_myt_today()
    today_date = date.fromisoformat(today_str)
    working_day = is_working_day(today_date)

    active_employees = await db.profiles.find(
        {"role": "employee", "status": "active"}, {"_id": 1}
    ).to_list(1000)
    total_employees = len(active_employees)

    attendance_records = await db.attendance.find(
        {"date": today_str}
    ).to_list(1000)

    present_today = len(attendance_records)
    still_working = sum(1 for r in attendance_records if r.get("status") == "working")
    auto_clocked_out = sum(1 for r in attendance_records if r.get("clock_out_reason") == "auto")
    absent_today = 0 if not working_day else (total_employees - present_today)

    return {
        "total_employees": total_employees,
        "present_today": present_today,
        "absent_today": absent_today,
        "still_working": still_working,
        "auto_clocked_out": auto_clocked_out,
        "date": today_str,
        "is_working_day": working_day
    }

# ─────────────────────────────────────────────
# Leave Requests
# ─────────────────────────────────────────────

@api_router.post("/leave-requests")
async def submit_leave_request(request: Request, current_user: dict = Depends(require_active)):
    if current_user["role"] != "employee":
        raise HTTPException(status_code=403, detail="Only employees can submit leave requests")
    
    body = await request.json()
    date_from = body.get("date_from")
    date_to = body.get("date_to")
    reason = body.get("reason", "").strip()
    
    if not date_from or not date_to:
        raise HTTPException(status_code=400, detail="date_from and date_to are required")
    if not reason:
        raise HTTPException(status_code=400, detail="reason is required")
    
    try:
        date.fromisoformat(date_from)
        date.fromisoformat(date_to)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    if date_from > date_to:
        raise HTTPException(status_code=400, detail="date_from must be before or equal to date_to")
    
    # Check for overlapping pending or approved requests
    existing = await db.leave_requests.find({
        "employee_id": current_user["id"],
        "status": {"$in": ["pending", "approved"]},
        "$or": [
            {"date_from": {"$lte": date_to}, "date_to": {"$gte": date_from}}
        ]
    }).to_list(10)
    
    if existing:
        raise HTTPException(status_code=409, detail="Leave request overlaps with an existing pending or approved request")
    
    leave_doc = {
        "employee_id": current_user["id"],
        "employee_name": current_user.get("full_name", ""),
        "date_from": date_from,
        "date_to": date_to,
        "reason": reason,
        "status": "pending",
        "requested_at": datetime.now(timezone.utc).isoformat(),
        "reviewed_at": None,
        "reviewed_by": None,
        "boss_remarks": None
    }
    
    result = await db.leave_requests.insert_one(leave_doc)
    leave_doc["id"] = str(result.inserted_id)
    leave_doc.pop("_id", None)
    
    return leave_doc

@api_router.get("/leave-requests/my")
async def get_my_leave_requests(current_user: dict = Depends(require_active)):
    requests = await db.leave_requests.find(
        {"employee_id": current_user["id"]}
    ).sort("date_from", -1).to_list(100)
    return [doc_to_dict(r) for r in requests]

@api_router.get("/leave-requests/pending")
async def get_pending_leave_requests(current_user: dict = Depends(require_admin)):
    requests = await db.leave_requests.find(
        {"status": "pending"}
    ).sort("requested_at", 1).to_list(100)
    return [doc_to_dict(r) for r in requests]

@api_router.put("/leave-requests/{request_id}/approve")
async def approve_leave_request(request_id: str, request: Request, current_user: dict = Depends(require_admin)):
    body = await request.json()
    boss_remarks = body.get("boss_remarks", "").strip()
    
    try:
        result = await db.leave_requests.update_one(
            {"_id": ObjectId(request_id), "status": "pending"},
            {"$set": {
                "status": "approved",
                "reviewed_at": datetime.now(timezone.utc).isoformat(),
                "reviewed_by": current_user["id"],
                "boss_remarks": boss_remarks if boss_remarks else None
            }}
        )
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid request ID")
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Pending leave request not found")
    
    updated = await db.leave_requests.find_one({"_id": ObjectId(request_id)})
    return doc_to_dict(updated)

@api_router.put("/leave-requests/{request_id}/reject")
async def reject_leave_request(request_id: str, request: Request, current_user: dict = Depends(require_admin)):
    body = await request.json()
    boss_remarks = body.get("boss_remarks", "").strip()
    
    try:
        result = await db.leave_requests.update_one(
            {"_id": ObjectId(request_id), "status": "pending"},
            {"$set": {
                "status": "rejected",
                "reviewed_at": datetime.now(timezone.utc).isoformat(),
                "reviewed_by": current_user["id"],
                "boss_remarks": boss_remarks if boss_remarks else None
            }}
        )
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid request ID")
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Pending leave request not found")
    
    updated = await db.leave_requests.find_one({"_id": ObjectId(request_id)})
    return doc_to_dict(updated)

@api_router.get("/leave-requests")
async def get_all_leave_requests(
    employee_id: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(require_admin)
):
    query = {}
    if employee_id:
        query["employee_id"] = employee_id
    if status:
        query["status"] = status
    if date_from or date_to:
        date_q = {}
        if date_from:
            date_q["$gte"] = date_from
        if date_to:
            date_q["$lte"] = date_to
        query["date_from"] = date_q
    
    requests = await db.leave_requests.find(query).sort("date_from", -1).to_list(500)
    return [doc_to_dict(r) for r in requests]

# ─────────────────────────────────────────────
# MEETINGS / APPOINTMENTS
# ─────────────────────────────────────────────

def _format_myt(dt: datetime) -> str:
    return dt.astimezone(MYT).strftime("%d %b %Y, %I:%M %p")

def _parse_myt_datetime(value: str) -> datetime:
    """Parse an ISO datetime string. Naive values are assumed to be in MYT."""
    dt = datetime.fromisoformat(value)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=MYT)
    return dt.astimezone(timezone.utc)

def _build_meeting_alert_text(meeting: dict, action: str = "New Meeting Scheduled") -> str:
    start_utc = datetime.fromisoformat(meeting["start_at"])
    lines = [
        f"📅 {action}",
        "",
        f"Employee: {meeting.get('employee_name', '')}",
        f"Meeting with: {meeting.get('meeting_with', '')}",
        f"Purpose: {meeting.get('purpose', '')}",
        f"Date & Time: {_format_myt(start_utc)}",
        f"Duration: {meeting.get('duration_minutes', 60)} mins",
        f"Location: {meeting.get('location') or 'Not specified'}"
    ]
    return "\n".join(lines)

async def _post_meeting_alert(message_text: str):
    try:
        channel = await db.channels.find_one({"name": "Meeting Alerts", "type": "public"})
        if not channel:
            return
        channel_id = str(channel["_id"])
        now_iso = datetime.now(timezone.utc).isoformat()
        msg_doc = {
            "channel_id": ObjectId(channel_id),
            "sender_id": "system",
            "sender_name": "Meeting Alerts",
            "sender_role": "system",
            "content": message_text,
            "attachment": None,
            "reactions": [],
            "created_at": now_iso,
            "updated_at": now_iso,
            "edited_at": None,
            "deleted": False,
            "deleted_at": None,
            "deleted_by": None
        }
        result = await db.messages.insert_one(msg_doc)
        msg_doc["id"] = str(result.inserted_id)
        msg_doc["channel_id"] = channel_id
        msg_doc.pop("_id", None)
        await ws_manager.broadcast(channel_id, {"type": "new_message", "message": msg_doc})
    except Exception as e:
        logger.error(f"Failed to post meeting alert: {e}")

@api_router.post("/meetings")
async def create_meeting(request: Request, current_user: dict = Depends(require_active)):
    body = await request.json()

    title = (body.get("title") or "").strip()
    meeting_with = (body.get("meeting_with") or "").strip()
    purpose = (body.get("purpose") or "").strip()
    start_at_raw = body.get("start_at")
    duration_minutes = body.get("duration_minutes", 60)
    location = (body.get("location") or "").strip() or None

    if not title:
        raise HTTPException(status_code=400, detail="Title is required")
    if not meeting_with:
        raise HTTPException(status_code=400, detail="Meeting with is required")
    if not purpose:
        raise HTTPException(status_code=400, detail="Purpose is required")
    if not start_at_raw:
        raise HTTPException(status_code=400, detail="Start time is required")

    try:
        duration_minutes = int(duration_minutes)
        if duration_minutes < 1:
            duration_minutes = 60
    except Exception:
        duration_minutes = 60

    try:
        start_utc = _parse_myt_datetime(start_at_raw)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid start_at datetime format")

    now_utc = datetime.now(timezone.utc)
    if start_utc < now_utc:
        raise HTTPException(status_code=400, detail="Meeting start time must be in the future")

    now_iso = datetime.now(timezone.utc).isoformat()
    meeting_doc = {
        "employee_id": current_user["id"],
        "employee_name": current_user.get("full_name", ""),
        "title": title,
        "meeting_with": meeting_with,
        "start_at": start_utc.isoformat(),
        "duration_minutes": duration_minutes,
        "location": location,
        "purpose": purpose,
        "status": "scheduled",
        "whatsapp_notified": False,
        "reminder_sent": False,
        "created_at": now_iso,
        "updated_at": now_iso
    }

    result = await db.meetings.insert_one(meeting_doc)
    inserted_id = str(result.inserted_id)
    meeting_doc["id"] = inserted_id

    # WhatsApp notification to boss — non-fatal
    try:
        alert_text = _build_meeting_alert_text(meeting_doc)
        send_whatsapp(alert_text)
        await db.meetings.update_one(
            {"_id": ObjectId(inserted_id)},
            {"$set": {"whatsapp_notified": True, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    except Exception as e:
        logger.error(f"WhatsApp notification failed for meeting {inserted_id}: {e}")

    # In-app alert in Meeting Alerts channel — non-fatal
    try:
        await _post_meeting_alert(_build_meeting_alert_text(meeting_doc))
    except Exception as e:
        logger.error(f"Meeting alert post failed for meeting {inserted_id}: {e}")

    # Re-fetch and serialize cleanly to avoid returning any ObjectId fields
    inserted = await db.meetings.find_one({"_id": result.inserted_id})
    return doc_to_dict(inserted)

@api_router.get("/meetings/my")
async def get_my_meetings(current_user: dict = Depends(require_active)):
    meetings = await db.meetings.find(
        {"employee_id": current_user["id"]}
    ).sort("start_at", -1).to_list(500)
    return [doc_to_dict(m) for m in meetings]

@api_router.get("/meetings/upcoming")
async def get_upcoming_meetings(current_user: dict = Depends(require_admin)):
    now_iso = datetime.now(timezone.utc).isoformat()
    meetings = await db.meetings.find(
        {"status": "scheduled", "start_at": {"$gte": now_iso}}
    ).sort("start_at", 1).limit(10).to_list(10)
    return [doc_to_dict(m) for m in meetings]

@api_router.get("/meetings")
async def list_meetings(
    employee_id: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(require_admin)
):
    query = {}
    if employee_id:
        query["employee_id"] = employee_id
    if status:
        query["status"] = status

    if date_from or date_to:
        date_q = {}
        if date_from:
            try:
                start_local = datetime.fromisoformat(date_from)
                if start_local.tzinfo is None:
                    start_local = start_local.replace(tzinfo=MYT)
                date_q["$gte"] = start_local.astimezone(timezone.utc).isoformat()
            except Exception:
                raise HTTPException(status_code=400, detail="Invalid date_from")
        if date_to:
            try:
                end_local = datetime.fromisoformat(date_to)
                if end_local.tzinfo is None:
                    end_local = end_local.replace(tzinfo=MYT)
                date_q["$lte"] = end_local.astimezone(timezone.utc).isoformat()
            except Exception:
                raise HTTPException(status_code=400, detail="Invalid date_to")
        query["start_at"] = date_q

    meetings = await db.meetings.find(query).sort("start_at", -1).to_list(500)
    return [doc_to_dict(m) for m in meetings]

@api_router.put("/meetings/{meeting_id}/cancel")
async def cancel_meeting(meeting_id: str, current_user: dict = Depends(require_active)):
    try:
        meeting = await db.meetings.find_one({"_id": ObjectId(meeting_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid meeting ID")

    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting not found")

    if current_user["role"] not in ["admin", "boss"] and meeting.get("employee_id") != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    if meeting.get("status") != "scheduled":
        raise HTTPException(status_code=400, detail="Only scheduled meetings can be cancelled")

    now_iso = datetime.now(timezone.utc).isoformat()
    await db.meetings.update_one(
        {"_id": ObjectId(meeting_id)},
        {"$set": {"status": "cancelled", "updated_at": now_iso}}
    )

    start_utc = datetime.fromisoformat(meeting["start_at"])
    message = (
        "❌ Meeting Cancelled\n\n"
        f"Employee: {meeting.get('employee_name', '')}\n"
        f"Meeting with: {meeting.get('meeting_with', '')}\n"
        f"Purpose: {meeting.get('purpose', '')}\n"
        f"Was scheduled for: {_format_myt(start_utc)}\n"
        f"Cancelled by: {current_user.get('full_name', '')}"
    )

    try:
        send_whatsapp(message)
    except Exception as e:
        logger.error(f"WhatsApp cancel notification failed: {e}")

    try:
        await _post_meeting_alert(message)
    except Exception as e:
        logger.error(f"Meeting Alerts cancel notification failed: {e}")

    updated = await db.meetings.find_one({"_id": ObjectId(meeting_id)})
    return doc_to_dict(updated)

@api_router.post("/cron/dispatch-reminders")
async def dispatch_reminders(request: Request):
    cron_secret = request.headers.get("X-Cron-Secret")
    if not cron_secret or cron_secret != CRON_SECRET:
        raise HTTPException(status_code=401, detail="Unauthorized")

    now = datetime.now(timezone.utc)
    window_start = (now + timedelta(minutes=105)).isoformat()  # now + 1h45m
    window_end = (now + timedelta(minutes=135)).isoformat()     # now + 2h15m
    stale_threshold = (now - timedelta(hours=24)).isoformat()

    # Clean up stale scheduled meetings that already started more than 24h ago
    stale_result = await db.meetings.update_many(
        {"status": "scheduled", "start_at": {"$lt": stale_threshold}},
        {"$set": {"status": "completed", "updated_at": now.isoformat()}}
    )

    # Find meetings due for a 2-hour reminder
    query = {
        "status": "scheduled",
        "reminder_sent": False,
        "start_at": {"$gte": window_start, "$lte": window_end}
    }
    upcoming = await db.meetings.find(query).sort("start_at", 1).to_list(500)

    dispatched = 0
    failed = 0

    for meeting in upcoming:
        meeting_id = str(meeting["_id"])
        start_utc = datetime.fromisoformat(meeting["start_at"])
        start_myt = start_utc.astimezone(MYT).strftime("%d %b %Y, %I:%M %p")
        message = (
            "⏰ Meeting Reminder\n\n"
            "In 2 hours:\n"
            f"Employee: {meeting.get('employee_name', '')}\n"
            f"Meeting with: {meeting.get('meeting_with', '')}\n"
            f"Purpose: {meeting.get('purpose', '')}\n"
            f"Time: {start_myt}\n"
            f"Location: {meeting.get('location') or 'Not specified'}"
        )

        # Always mark reminder_sent=True so we don't retry on failures
        await db.meetings.update_one(
            {"_id": ObjectId(meeting_id)},
            {"$set": {"reminder_sent": True, "updated_at": now.isoformat()}}
        )

        try:
            send_whatsapp(message)
        except Exception as e:
            logger.error(f"WhatsApp reminder failed for meeting {meeting_id}: {e}")
            failed += 1

        try:
            await _post_meeting_alert(message)
        except Exception as e:
            logger.error(f"In-app reminder alert failed for meeting {meeting_id}: {e}")

        dispatched += 1

    return {
        "dispatched": dispatched,
        "failed": failed,
        "stale_completed": stale_result.modified_count
    }

# ─────────────────────────────────────────────
# Monthly Attendance Reports
# ─────────────────────────────────────────────

@api_router.get("/attendance/monthly-report")
async def get_monthly_attendance_report(
    year: int,
    month: int,
    current_user: dict = Depends(require_admin)
):
    """Get aggregated monthly attendance report for all employees."""
    await process_auto_clock_outs()
    
    # Get all active employees
    active_employees = await db.profiles.find(
        {"role": "employee", "status": "active"}
    ).to_list(1000)
    
    # Calculate working days in the month
    first_day = date(year, month, 1)
    if month == 12:
        last_day = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, month + 1, 1) - timedelta(days=1)
    
    working_days = []
    current = first_day
    while current <= last_day:
        if is_working_day(current):
            working_days.append(current)
        current += timedelta(days=1)
    
    working_days_count = len(working_days)
    
    # Get approved leave requests for the month
    leave_requests = await db.leave_requests.find({
        "status": "approved",
        "$or": [
            {"date_from": {"$lte": last_day.isoformat()}, "date_to": {"$gte": first_day.isoformat()}}
        ]
    }).to_list(500)
    
    # Build leave map: employee_id -> set of leave dates
    leave_map = {}
    for lr in leave_requests:
        emp_id = lr["employee_id"]
        if emp_id not in leave_map:
            leave_map[emp_id] = set()
        lr_start = date.fromisoformat(lr["date_from"])
        lr_end = date.fromisoformat(lr["date_to"])
        current = max(lr_start, first_day)
        while current <= min(lr_end, last_day):
            leave_map[emp_id].add(current)
            current += timedelta(days=1)
    
    # Get attendance records for the month
    attendance_records = await db.attendance.find({
        "date": {"$gte": first_day.isoformat(), "$lte": last_day.isoformat()}
    }).to_list(5000)
    
    # Build attendance map: (employee_id, date) -> record
    attendance_map = {}
    for att in attendance_records:
        key = (att["employee_id"], att["date"])
        attendance_map[key] = att
    
    # Calculate metrics for each employee
    threshold_time = time(9, LATE_ARRIVAL_THRESHOLD_MINUTES, 0)
    
    results = []
    for emp in active_employees:
        emp_id = str(emp["_id"])
        emp_name = emp.get("full_name", "")
        department = emp.get("department", "")
        
        days_present = 0
        days_on_leave = 0
        days_absent = 0
        total_working_minutes = 0
        late_arrivals_count = 0
        auto_clock_outs_count = 0
        
        for work_day in working_days:
            day_str = work_day.isoformat()
            att = attendance_map.get((emp_id, day_str))
            
            if att:
                days_present += 1
                total_working_minutes += att.get("working_duration_minutes", 0)
                
                # Check late arrival
                try:
                    clock_in_str = att.get("clock_in", "")
                    if clock_in_str:
                        clock_in_time = datetime.strptime(clock_in_str, "%H:%M:%S").time()
                        if clock_in_time > threshold_time:
                            late_arrivals_count += 1
                except (ValueError, TypeError):
                    pass
                
                # Check auto clock-out
                if att.get("clock_out_reason") == "auto":
                    auto_clock_outs_count += 1
            elif emp_id in leave_map and work_day in leave_map[emp_id]:
                days_on_leave += 1
            else:
                days_absent += 1
        
        average_hours_per_day = 0
        if days_present > 0:
            average_hours_per_day = round(total_working_minutes / days_present / 60, 2)
        
        results.append({
            "employee_id": emp_id,
            "employee_name": emp_name,
            "department": department,
            "working_days_in_month": working_days_count,
            "days_present": days_present,
            "days_on_leave": days_on_leave,
            "days_absent": days_absent,
            "total_working_minutes": total_working_minutes,
            "average_hours_per_day": average_hours_per_day,
            "late_arrivals_count": late_arrivals_count,
            "auto_clock_outs_count": auto_clock_outs_count
        })
    
    return results

@api_router.get("/attendance/monthly-report/{employee_id}")
async def get_employee_monthly_drilldown(
    employee_id: str,
    year: int,
    month: int,
    current_user: dict = Depends(require_admin)
):
    """Get day-by-day attendance breakdown for an employee in a month."""
    await process_auto_clock_outs()
    
    # Verify employee exists
    try:
        employee = await db.profiles.find_one({"_id": ObjectId(employee_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid employee ID")
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Calculate date range
    first_day = date(year, month, 1)
    if month == 12:
        last_day = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, month + 1, 1) - timedelta(days=1)
    
    # Get working days
    working_days = []
    current = first_day
    while current <= last_day:
        if is_working_day(current):
            working_days.append(current)
        current += timedelta(days=1)
    
    working_days_count = len(working_days)
    
    # Get approved leave requests
    leave_requests = await db.leave_requests.find({
        "employee_id": employee_id,
        "status": "approved",
        "$or": [
            {"date_from": {"$lte": last_day.isoformat()}, "date_to": {"$gte": first_day.isoformat()}}
        ]
    }).to_list(100)
    
    # Build leave date set
    leave_dates = set()
    for lr in leave_requests:
        lr_start = date.fromisoformat(lr["date_from"])
        lr_end = date.fromisoformat(lr["date_to"])
        current = max(lr_start, first_day)
        while current <= min(lr_end, last_day):
            leave_dates.add(current)
            current += timedelta(days=1)
    
    # Get attendance records
    attendance_records = await db.attendance.find({
        "employee_id": employee_id,
        "date": {"$gte": first_day.isoformat(), "$lte": last_day.isoformat()}
    }).to_list(100)
    
    attendance_map = {att["date"]: att for att in attendance_records}
    
    # Build daily breakdown
    threshold_time = time(9, LATE_ARRIVAL_THRESHOLD_MINUTES, 0)
    
    days_present = 0
    days_on_leave = 0
    days_absent = 0
    total_working_minutes = 0
    late_arrivals_count = 0
    auto_clock_outs_count = 0
    
    daily_breakdown = []
    current = first_day
    while current <= last_day:
        day_str = current.isoformat()
        
        if not is_working_day(current):
            daily_breakdown.append({
                "date": day_str,
                "status": "weekend",
                "clock_in": None,
                "clock_out": None,
                "working_duration_minutes": None,
                "is_late": False,
                "is_auto_clock_out": False
            })
        else:
            att = attendance_map.get(day_str)
            if att:
                days_present += 1
                total_working_minutes += att.get("working_duration_minutes", 0)
                
                # Check late arrival
                is_late = False
                try:
                    clock_in_str = att.get("clock_in", "")
                    if clock_in_str:
                        clock_in_time = datetime.strptime(clock_in_str, "%H:%M:%S").time()
                        if clock_in_time > threshold_time:
                            is_late = True
                            late_arrivals_count += 1
                except (ValueError, TypeError):
                    pass
                
                # Check auto clock-out
                is_auto = att.get("clock_out_reason") == "auto"
                if is_auto:
                    auto_clock_outs_count += 1
                
                daily_breakdown.append({
                    "date": day_str,
                    "status": "present",
                    "clock_in": att.get("clock_in"),
                    "clock_out": att.get("clock_out"),
                    "working_duration_minutes": att.get("working_duration_minutes"),
                    "is_late": is_late,
                    "is_auto_clock_out": is_auto
                })
            elif current in leave_dates:
                days_on_leave += 1
                daily_breakdown.append({
                    "date": day_str,
                    "status": "on_leave",
                    "clock_in": None,
                    "clock_out": None,
                    "working_duration_minutes": None,
                    "is_late": False,
                    "is_auto_clock_out": False
                })
            else:
                days_absent += 1
                daily_breakdown.append({
                    "date": day_str,
                    "status": "absent",
                    "clock_in": None,
                    "clock_out": None,
                    "working_duration_minutes": None,
                    "is_late": False,
                    "is_auto_clock_out": False
                })
        
        current += timedelta(days=1)
    
    average_hours_per_day = 0
    if days_present > 0:
        average_hours_per_day = round(total_working_minutes / days_present / 60, 2)
    
    return {
        "employee_id": employee_id,
        "employee_name": employee.get("full_name", ""),
        "department": employee.get("department", ""),
        "year": year,
        "month": month,
        "working_days_in_month": working_days_count,
        "days_present": days_present,
        "days_on_leave": days_on_leave,
        "days_absent": days_absent,
        "total_working_minutes": total_working_minutes,
        "average_hours_per_day": average_hours_per_day,
        "late_arrivals_count": late_arrivals_count,
        "auto_clock_outs_count": auto_clock_outs_count,
        "daily_breakdown": daily_breakdown
    }

# ─────────────────────────────────────────────
# Misc
# ─────────────────────────────────────────────

@api_router.get("/attendance/history")
async def get_attendance_history(
    start_date: str,
    end_date: str,
    employee_id: Optional[str] = None,
    current_user: dict = Depends(require_admin)
):
    try:
        query = {"date": {"$gte": start_date, "$lte": end_date}}
        if employee_id:
            query["employee_id"] = employee_id

        records = await db.attendance.find(query).sort("date", -1).to_list(2000)

        emp_ids = list({str(r.get("employee_id")) for r in records})
        emp_map = {}
        if emp_ids:
            from bson import ObjectId as _OID
            oid_list = []
            for eid in emp_ids:
                try:
                    oid_list.append(_OID(eid))
                except Exception:
                    pass
            if oid_list:
                employees = await db.profiles.find(
                    {"_id": {"$in": oid_list}}, {"password_hash": 0}
                ).to_list(1000)
                for emp in employees:
                    emp_map[str(emp["_id"])] = emp

        result = []
        for r in records:
            emp = emp_map.get(str(r.get("employee_id")), {})
            is_late = None
            if r.get("clock_in_at"):
                try:
                    cin = datetime.fromisoformat(r["clock_in_at"])
                    late_threshold = cin.replace(hour=9, minute=15, second=0, microsecond=0)
                    is_late = cin > late_threshold
                except (ValueError, TypeError):
                    is_late = None

            result.append({
                "employee_id": str(r.get("employee_id", "")),
                "employee_name": emp.get("full_name", r.get("employee_name", "")),
                "department": emp.get("department", r.get("department", "")),
                "date": r.get("date", ""),
                "clock_in": r.get("clock_in"),
                "clock_out": r.get("clock_out"),
                "clock_in_at": r.get("clock_in_at"),
                "clock_out_at": r.get("clock_out_at"),
                "working_duration_minutes": r.get("working_duration_minutes"),
                "working_duration_display": r.get("working_duration_display", ""),
                "clock_out_reason": r.get("clock_out_reason"),
                "status": r.get("status", ""),
                "is_late": is_late
            })

        return result
    except Exception as e:
        logger.error(f"Error in get_attendance_history: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/attendance/history/summary")
async def get_attendance_history_summary(
    start_date: str,
    end_date: str,
    employee_id: Optional[str] = None,
    current_user: dict = Depends(require_admin)
):
    try:
        query = {"date": {"$gte": start_date, "$lte": end_date}}
        if employee_id:
            query["employee_id"] = employee_id

        records = await db.attendance.find(query).to_list(2000)

        present_days = len(records)
        auto_clock_out_count = sum(1 for r in records if r.get("clock_out_reason") == "auto")
        total_minutes = sum(r.get("working_duration_minutes", 0) or 0 for r in records)

        late_count = 0
        for r in records:
            if r.get("clock_in_at"):
                try:
                    cin = datetime.fromisoformat(r["clock_in_at"])
                    late_threshold = cin.replace(hour=9, minute=15, second=0, microsecond=0)
                    if cin > late_threshold:
                        late_count += 1
                except (ValueError, TypeError):
                    pass

        start_d = date.fromisoformat(start_date)
        end_d = date.fromisoformat(end_date)
        total_working_days = 0
        cur = start_d
        while cur <= end_d:
            if is_working_day(cur):
                total_working_days += 1
            cur += timedelta(days=1)

        if employee_id:
            absent_days = max(0, total_working_days - present_days)
        else:
            active_count = await db.profiles.count_documents({"role": "employee", "status": "active"})
            absent_days = max(0, total_working_days * active_count - present_days)

        avg_hours = round(total_minutes / 60 / present_days, 1) if present_days > 0 else 0

        return {
            "total_days": total_working_days,
            "present_days": present_days,
            "absent_days": absent_days,
            "late_count": late_count,
            "auto_clock_out_count": auto_clock_out_count,
            "avg_working_hours": avg_hours,
            "total_working_hours": round(total_minutes / 60, 1)
        }
    except Exception as e:
        logger.error(f"Error in get_attendance_history_summary: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/attendance/my-history")
async def get_my_attendance_history(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(require_active)
):
    try:
        today_str = get_myt_today()
        today_d = date.fromisoformat(today_str)
        first_of_month = today_d.replace(day=1)

        sd = start_date or first_of_month.isoformat()
        ed = end_date or today_str

        records = await db.attendance.find({
            "employee_id": current_user["id"],
            "date": {"$gte": sd, "$lte": ed}
        }).sort("date", -1).to_list(200)

        result = []
        for r in records:
            is_late = None
            if r.get("clock_in_at"):
                try:
                    cin = datetime.fromisoformat(r["clock_in_at"])
                    late_threshold = cin.replace(hour=9, minute=15, second=0, microsecond=0)
                    is_late = cin > late_threshold
                except (ValueError, TypeError):
                    is_late = None

            result.append({
                "date": r.get("date", ""),
                "clock_in": r.get("clock_in"),
                "clock_out": r.get("clock_out"),
                "working_duration_display": r.get("working_duration_display", ""),
                "status": r.get("status", ""),
                "clock_out_reason": r.get("clock_out_reason"),
                "is_late": is_late
            })

        present_days = len(records)
        late_count = sum(1 for r in result if r["is_late"])
        total_minutes = sum(
            (rec.get("working_duration_minutes") or 0)
            for rec in records
        )
        avg_hours = round(total_minutes / 60 / present_days, 1) if present_days > 0 else 0

        total_working_days = 0
        cur = date.fromisoformat(sd)
        end_d = date.fromisoformat(ed)
        while cur <= end_d:
            if is_working_day(cur):
                total_working_days += 1
            cur += timedelta(days=1)
        absent_days = max(0, total_working_days - present_days)

        return {
            "records": result,
            "summary": {
                "present_days": present_days,
                "absent_days": absent_days,
                "late_count": late_count,
                "avg_working_hours": avg_hours
            }
        }
    except Exception as e:
        logger.error(f"Error in get_my_attendance_history: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/attendance/employee-summary/{employee_id}")
async def get_employee_attendance_summary(employee_id: str, current_user: dict = Depends(require_admin)):
    try:
        today_str = get_myt_today()
        today_d = date.fromisoformat(today_str)

        this_month_start = today_d.replace(day=1).isoformat()
        last_month_end = (today_d.replace(day=1) - timedelta(days=1)).isoformat()
        last_month_start = (today_d.replace(day=1) - timedelta(days=1)).replace(day=1).isoformat()

        async def get_month_summary(sd, ed):
            records = await db.attendance.find({
                "employee_id": employee_id,
                "date": {"$gte": sd, "$lte": ed}
            }).to_list(200)

            present = len(records)
            late = 0
            total_min = 0
            auto_count = 0
            for r in records:
                if r.get("clock_in_at"):
                    try:
                        cin = datetime.fromisoformat(r["clock_in_at"])
                        if cin > cin.replace(hour=9, minute=15, second=0, microsecond=0):
                            late += 1
                    except (ValueError, TypeError):
                        pass
                total_min += r.get("working_duration_minutes", 0) or 0
                if r.get("clock_out_reason") == "auto":
                    auto_count += 1

            start_d = date.fromisoformat(sd)
            end_d = date.fromisoformat(ed)
            wd = 0
            cur = start_d
            while cur <= end_d:
                if is_working_day(cur):
                    wd += 1
                cur += timedelta(days=1)

            present_pct = (present / wd * 100) if wd > 0 else 0
            avg_h = round(total_min / 60 / present, 1) if present > 0 else 0

            if present_pct > 95 and late == 0:
                rating = "Excellent"
            elif present_pct > 90 and late < 3:
                rating = "Good"
            else:
                rating = "Needs Improvement"

            return {
                "present_days": present,
                "absent_days": max(0, wd - present),
                "late_count": late,
                "avg_working_hours": avg_h,
                "auto_clock_out_count": auto_count,
                "present_percentage": round(present_pct, 1),
                "rating": rating
            }

        this_month = await get_month_summary(this_month_start, today_str)
        last_month = await get_month_summary(last_month_start, last_month_end)

        daily_records = await db.attendance.find({
            "employee_id": employee_id,
            "date": {"$gte": this_month_start, "$lte": today_str}
        }).sort("date", 1).to_list(100)

        daily_hours = []
        for r in daily_records:
            mins = r.get("working_duration_minutes", 0) or 0
            daily_hours.append({
                "date": r.get("date", ""),
                "hours": round(mins / 60, 1)
            })

        return {
            "this_month": this_month,
            "last_month": last_month,
            "daily_hours": daily_hours
        }
    except Exception as e:
        logger.error(f"Error in get_employee_attendance_summary: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ─────────────────────────────────────────────
# MESSAGING ROUTES
# ─────────────────────────────────────────────

@api_router.get("/channels")
async def list_channels(current_user: dict = Depends(require_active)):
    user_id = current_user["id"]
    query = {
        "$or": [
            {"type": "public"},
            {"type": {"$in": ["private", "dm"]}, "members": user_id}
        ]
    }
    channels = await db.channels.find(query).sort("name", 1).to_list(100)
    result = []
    for ch in channels:
        ch_dict = doc_to_dict(ch)
        if ch_dict["type"] == "dm":
            other_id = next((m for m in ch_dict.get("members", []) if m != user_id), None)
            if other_id:
                other = await db.profiles.find_one({"_id": ObjectId(other_id)}, {"password_hash": 0})
                if other:
                    ch_dict["other_user"] = {
                        "id": str(other["_id"]),
                        "full_name": other.get("full_name", ""),
                        "role": other.get("role", ""),
                        "status": other.get("status", "")
                    }
        result.append(ch_dict)
    return result

@api_router.post("/channels")
async def create_channel(request: Request, current_user: dict = Depends(require_admin)):
    body = await request.json()
    name = body.get("name", "").strip()
    channel_type = body.get("type", "")
    members = body.get("members", [])

    if not name:
        raise HTTPException(status_code=400, detail="Channel name is required")
    if channel_type not in ["public", "private"]:
        raise HTTPException(status_code=400, detail="Type must be 'public' or 'private'")

    existing = await db.channels.find_one({"name": name})
    if existing:
        raise HTTPException(status_code=409, detail="Channel name already exists")

    channel_doc = {
        "name": name,
        "type": channel_type,
        "members": members if channel_type == "private" else [],
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.channels.insert_one(channel_doc)
    channel_doc["id"] = str(result.inserted_id)
    channel_doc.pop("_id", None)
    return channel_doc

@api_router.get("/channels/unread-counts")
async def get_unread_counts(current_user: dict = Depends(require_active)):
    user_id = current_user["id"]
    query = {
        "$or": [
            {"type": "public"},
            {"type": {"$in": ["private", "dm"]}, "members": user_id}
        ]
    }
    channels = await db.channels.find(query).to_list(100)
    result = {}
    for ch in channels:
        ch_id = str(ch["_id"])
        read_record = await db.message_reads.find_one({
            "user_id": user_id,
            "channel_id": ch["_id"]
        })
        if read_record and read_record.get("last_read_at"):
            unread_count = await db.messages.count_documents({
                "channel_id": ch["_id"],
                "deleted": False,
                "created_at": {"$gt": read_record["last_read_at"]},
                "sender_id": {"$ne": user_id}
            })
        else:
            unread_count = await db.messages.count_documents({
                "channel_id": ch["_id"],
                "deleted": False,
                "sender_id": {"$ne": user_id}
            })
        result[ch_id] = unread_count
    return result

@api_router.get("/channels/{channel_id}/messages")
async def get_channel_messages(
    channel_id: str,
    before: Optional[str] = None,
    current_user: dict = Depends(require_active)
):
    try:
        ch = await db.channels.find_one({"_id": ObjectId(channel_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid channel ID")
    if not ch:
        raise HTTPException(status_code=404, detail="Channel not found")

    if ch["type"] != "public" and current_user["id"] not in ch.get("members", []):
        raise HTTPException(status_code=403, detail="Access denied to this channel")

    query = {"channel_id": ObjectId(channel_id)}
    if before:
        try:
            before_dt = datetime.fromisoformat(before)
            query["created_at"] = {"$lt": before_dt.isoformat()}
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid 'before' timestamp")

    messages = await db.messages.find(query).sort("created_at", -1).limit(50).to_list(50)
    result = []
    for msg in reversed(messages):
        msg_dict = doc_to_dict(msg)
        msg_dict["channel_id"] = str(msg_dict["channel_id"])
        result.append(msg_dict)
    return result

@api_router.post("/channels/{channel_id}/messages")
async def send_channel_message(
    channel_id: str,
    request: Request,
    current_user: dict = Depends(require_active)
):
    try:
        ch = await db.channels.find_one({"_id": ObjectId(channel_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid channel ID")
    if not ch:
        raise HTTPException(status_code=404, detail="Channel not found")

    if ch["type"] != "public" and current_user["id"] not in ch.get("members", []):
        raise HTTPException(status_code=403, detail="Access denied to this channel")

    body = await request.json()
    content = body.get("content", "").strip()
    attachment = body.get("attachment")
    if not content and not attachment:
        raise HTTPException(status_code=400, detail="Message content or attachment is required")
    if len(content) > 2000:
        raise HTTPException(status_code=400, detail="Message must be 2000 characters or less")

    now_iso = datetime.now(timezone.utc).isoformat()
    msg_doc = {
        "channel_id": ObjectId(channel_id),
        "sender_id": current_user["id"],
        "sender_name": current_user.get("full_name", ""),
        "sender_role": current_user.get("role", ""),
        "content": content,
        "attachment": attachment,
        "reactions": [],
        "created_at": now_iso,
        "edited_at": None,
        "deleted": False,
        "deleted_at": None,
        "deleted_by": None
    }
    result = await db.messages.insert_one(msg_doc)
    msg_doc["id"] = str(result.inserted_id)
    msg_doc["channel_id"] = str(msg_doc["channel_id"])
    msg_doc.pop("_id", None)
    return msg_doc

@api_router.put("/messages/{message_id}")
async def edit_message(
    message_id: str,
    request: Request,
    current_user: dict = Depends(require_active)
):
    body = await request.json()
    content = body.get("content", "").strip()
    if not content:
        raise HTTPException(status_code=400, detail="Message content is required")
    if len(content) > 2000:
        raise HTTPException(status_code=400, detail="Message must be 2000 characters or less")

    try:
        msg = await db.messages.find_one({"_id": ObjectId(message_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid message ID")
    if not msg:
        raise HTTPException(status_code=404, detail="Message not found")
    if msg["sender_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Only the sender can edit this message")
    if msg.get("deleted"):
        raise HTTPException(status_code=400, detail="Cannot edit a deleted message")

    now_iso = datetime.now(timezone.utc).isoformat()
    await db.messages.update_one(
        {"_id": ObjectId(message_id)},
        {"$set": {"content": content, "edited_at": now_iso}}
    )
    updated = await db.messages.find_one({"_id": ObjectId(message_id)})
    updated_dict = doc_to_dict(updated)
    updated_dict["channel_id"] = str(updated_dict["channel_id"])
    await ws_manager.broadcast(str(updated["channel_id"]), {
        "type": "message_edited",
        "message": updated_dict
    })
    return updated_dict

@api_router.delete("/messages/{message_id}")
async def delete_message(
    message_id: str,
    current_user: dict = Depends(require_active)
):
    try:
        msg = await db.messages.find_one({"_id": ObjectId(message_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid message ID")
    if not msg:
        raise HTTPException(status_code=404, detail="Message not found")
    if msg["sender_id"] != current_user["id"] and current_user["role"] not in ["admin", "boss"]:
        raise HTTPException(status_code=403, detail="Only the sender or an admin can delete this message")
    if msg.get("deleted"):
        raise HTTPException(status_code=400, detail="Message is already deleted")

    now_iso = datetime.now(timezone.utc).isoformat()
    await db.messages.update_one(
        {"_id": ObjectId(message_id)},
        {"$set": {"deleted": True, "deleted_at": now_iso, "deleted_by": current_user["id"], "content": None}}
    )
    updated = await db.messages.find_one({"_id": ObjectId(message_id)})
    await ws_manager.broadcast(str(updated["channel_id"]), {
        "type": "message_deleted",
        "message_id": message_id,
        "channel_id": str(updated["channel_id"])
    })
    return {"message": "Message deleted"}

@api_router.post("/messages/{message_id}/reactions")
async def toggle_reaction(
    message_id: str,
    request: Request,
    current_user: dict = Depends(require_active)
):
    body = await request.json()
    emoji = body.get("emoji", "").strip()
    if not emoji:
        raise HTTPException(status_code=400, detail="Emoji is required")

    try:
        oid = ObjectId(message_id)
        msg = await db.messages.find_one({"_id": oid})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid message ID")
    if not msg:
        raise HTTPException(status_code=404, detail="Message not found")

    user_id = current_user["id"]

    # Attempt toggle-off: atomically pull user_id from the matching emoji's users array
    pull_result = await db.messages.update_one(
        {"_id": oid, "reactions": {"$elemMatch": {"emoji": emoji, "users": user_id}}},
        {"$pull": {"reactions.$.users": user_id}}
    )
    if pull_result.matched_count > 0:
        # Clean up any reaction entries left with an empty users array
        await db.messages.update_one(
            {"_id": oid},
            {"$pull": {"reactions": {"users": {"$size": 0}}}}
        )
    else:
        # Toggle on: add user_id to existing emoji entry, or create a new one
        add_result = await db.messages.update_one(
            {"_id": oid, "reactions.emoji": emoji},
            {"$addToSet": {"reactions.$.users": user_id}}
        )
        if add_result.matched_count == 0:
            await db.messages.update_one(
                {"_id": oid, "reactions.emoji": {"$ne": emoji}},
                {"$push": {"reactions": {"emoji": emoji, "users": [user_id]}}}
            )

    updated = await db.messages.find_one({"_id": oid})
    updated_reactions = updated.get("reactions", [])
    channel_id = str(updated["channel_id"])

    await ws_manager.broadcast(channel_id, {
        "type": "reaction_update",
        "message_id": message_id,
        "channel_id": channel_id,
        "reactions": updated_reactions
    })

    return {"reactions": updated_reactions}

@api_router.post("/chat/upload")
async def upload_chat_file(
    request: Request,
    file: UploadFile = File(...),
    channel_id: str = Form(...),
    current_user: dict = Depends(require_active)
):
    try:
        ch = await db.channels.find_one({"_id": ObjectId(channel_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid channel ID")
    if not ch:
        raise HTTPException(status_code=404, detail="Channel not found")
    if ch["type"] != "public" and current_user["id"] not in ch.get("members", []):
        raise HTTPException(status_code=403, detail="Access denied to this channel")

    file_bytes = await file.read()
    if len(file_bytes) > 52428800:
        raise HTTPException(status_code=400, detail="File size exceeds 50MB limit")

    storage_path = f"{APP_STORAGE_PREFIX}/chat/{channel_id}/{uuid.uuid4().hex}_{file.filename}"
    try:
        put_object(storage_path, file_bytes, file.content_type or "application/octet-stream", timeout=300)
    except Exception as exc:
        raise HTTPException(status_code=500, detail="Failed to upload file")

    return {
        "url": f"{request.base_url}api/files/{storage_path}",
        "filename": file.filename,
        "file_size": len(file_bytes),
        "mime_type": file.content_type or "application/octet-stream",
        "storage_path": storage_path,
        "is_gif": False
    }

@api_router.post("/channels/{channel_id}/read")
async def mark_channel_read(
    channel_id: str,
    current_user: dict = Depends(require_active)
):
    try:
        ch = await db.channels.find_one({"_id": ObjectId(channel_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid channel ID")
    if not ch:
        raise HTTPException(status_code=404, detail="Channel not found")

    if ch["type"] != "public" and current_user["id"] not in ch.get("members", []):
        raise HTTPException(status_code=403, detail="Access denied to this channel")

    last_msg = await db.messages.find(
        {"channel_id": ObjectId(channel_id), "deleted": False}
    ).sort("created_at", -1).limit(1).to_list(1)

    last_msg_id = last_msg[0]["_id"] if last_msg else None
    now_iso = datetime.now(timezone.utc).isoformat()

    await db.message_reads.update_one(
        {"user_id": current_user["id"], "channel_id": ObjectId(channel_id)},
        {"$set": {
            "last_read_message_id": last_msg_id,
            "last_read_at": now_iso
        }},
        upsert=True
    )
    return {"message": "Channel marked as read"}

@api_router.post("/dms")
async def create_or_get_dm(request: Request, current_user: dict = Depends(require_active)):
    try:
        body = await request.json()
        target_user_id = (body.get("user_id") or "").strip()
        logger.info(f"POST /api/dms: target_user_id={target_user_id}, current_user={current_user['id']}")
        if not target_user_id:
            raise HTTPException(status_code=400, detail="user_id is required")
        if target_user_id == current_user["id"]:
            raise HTTPException(status_code=400, detail="Cannot start a DM with yourself")

        try:
            target = await db.profiles.find_one({"_id": ObjectId(target_user_id)}, {"password_hash": 0})
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid user ID")
        if not target:
            raise HTTPException(status_code=404, detail="User not found")
        if target.get("status") != "active":
            raise HTTPException(status_code=400, detail="User is not active")

        dm_between = sorted([current_user["id"], target_user_id])

        existing = await db.channels.find_one({
            "type": "dm",
            "$or": [
                {"dm_between": dm_between},
                {"members": {"$all": [current_user["id"], target_user_id], "$size": 2}}
            ]
        })
        if existing:
            logger.info(f"POST /api/dms: existing DM found channel_id={existing.get('_id')}")
            ch_dict = doc_to_dict(existing)
            ch_dict["other_user"] = {
                "id": str(target["_id"]),
                "full_name": target.get("full_name", ""),
                "role": target.get("role", ""),
                "status": target.get("status", "")
            }
            return ch_dict

        now_iso = datetime.now(timezone.utc).isoformat()
        dm_doc = {
            "type": "dm",
            "members": [current_user["id"], target_user_id],
            "dm_between": dm_between,
            "created_by": current_user["id"],
            "created_at": now_iso
        }
        result = await db.channels.insert_one(dm_doc)
        logger.info(f"POST /api/dms: created new DM channel_id={result.inserted_id}")
        dm_doc["id"] = str(result.inserted_id)
        dm_doc.pop("_id", None)
        dm_doc["other_user"] = {
            "id": str(target["_id"]),
            "full_name": target.get("full_name", ""),
            "role": target.get("role", ""),
            "status": target.get("status", "")
        }
        return dm_doc
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"POST /dms unhandled error: {e}")
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/dms")
async def list_dms(current_user: dict = Depends(require_active)):
    user_id = current_user["id"]
    dms = await db.channels.find({"type": "dm", "members": user_id}).sort("created_at", -1).to_list(100)
    result = []
    for dm in dms:
        dm_dict = doc_to_dict(dm)
        other_id = next((m for m in dm_dict.get("members", []) if m != user_id), None)
        if other_id:
            other = await db.profiles.find_one({"_id": ObjectId(other_id)}, {"password_hash": 0})
            if other:
                dm_dict["other_user"] = {
                    "id": str(other["_id"]),
                    "full_name": other.get("full_name", ""),
                    "role": other.get("role", ""),
                    "status": other.get("status", "")
                }
        result.append(dm_dict)
    return result

@api_router.post("/channels/{channel_id}/members")
async def add_channel_member(
    channel_id: str,
    request: Request,
    current_user: dict = Depends(require_admin)
):
    body = await request.json()
    target_user_id = body.get("user_id", "").strip()
    if not target_user_id:
        raise HTTPException(status_code=400, detail="user_id is required")

    try:
        ch = await db.channels.find_one({"_id": ObjectId(channel_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid channel ID")
    if not ch:
        raise HTTPException(status_code=404, detail="Channel not found")
    if ch["type"] != "private":
        raise HTTPException(status_code=400, detail="Member management only applies to private channels")

    try:
        target = await db.profiles.find_one({"_id": ObjectId(target_user_id)}, {"password_hash": 0})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid user ID")
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    if target.get("status") != "active":
        raise HTTPException(status_code=400, detail="User is not active")

    await db.channels.update_one(
        {"_id": ObjectId(channel_id)},
        {"$addToSet": {"members": target_user_id}}
    )
    return {"message": "Member added to channel"}

@api_router.delete("/channels/{channel_id}/members/{user_id}")
async def remove_channel_member(
    channel_id: str,
    user_id: str,
    current_user: dict = Depends(require_admin)
):
    try:
        ch = await db.channels.find_one({"_id": ObjectId(channel_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid channel ID")
    if not ch:
        raise HTTPException(status_code=404, detail="Channel not found")
    if ch["type"] != "private":
        raise HTTPException(status_code=400, detail="Member management only applies to private channels")

    await db.channels.update_one(
        {"_id": ObjectId(channel_id)},
        {"$pull": {"members": user_id}}
    )
    return {"message": "Member removed from channel"}

@api_router.get("/")
async def root():
    return {"message": "Performance Pulse API", "status": "ok"}

app.include_router(api_router)

# ─────────────────────────────────────────────
# WEBSOCKET CHAT
# ─────────────────────────────────────────────
# Render free tier: the server sleeps after 15 min of inactivity.
# When the server wakes, WS connections are broken. The frontend must
# auto-reconnect with exponential backoff and re-fetch unread counts
# via REST on every successful reconnect, since messages may have
# arrived while the WS was disconnected.

class ConnectionManager:
    def __init__(self):
        self.rooms: dict[str, set] = {}

    async def connect(self, channel_id: str, websocket: WebSocket):
        if channel_id not in self.rooms:
            self.rooms[channel_id] = set()
        self.rooms[channel_id].add(websocket)

    def disconnect(self, channel_id: str, websocket: WebSocket):
        if channel_id in self.rooms:
            self.rooms[channel_id].discard(websocket)
            if not self.rooms[channel_id]:
                del self.rooms[channel_id]

    async def broadcast(self, channel_id: str, message: dict, exclude: WebSocket = None):
        if channel_id not in self.rooms:
            return
        dead = []
        for ws in self.rooms[channel_id]:
            if ws is exclude:
                continue
            try:
                await ws.send_json(message)
            except Exception:
                dead.append(ws)
        for ws in dead:
            self.rooms[channel_id].discard(ws)
        if channel_id in self.rooms and not self.rooms[channel_id]:
            del self.rooms[channel_id]

ws_manager = ConnectionManager()

async def ws_authenticate(websocket: WebSocket) -> dict:
    token = websocket.cookies.get("access_token")
    if not token:
        await websocket.close(code=4001)
        return None
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            await websocket.close(code=4001)
            return None
        user = await db.profiles.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            await websocket.close(code=4001)
            return None
        result = doc_to_dict(user)
        result.pop("password_hash", None)
        if result.get("status") != "active":
            await websocket.close(code=4001)
            return None
        return result
    except Exception:
        await websocket.close(code=4001)
        return None

async def get_user_channel_ids(user_id: str) -> list:
    query = {
        "$or": [
            {"type": "public"},
            {"type": {"$in": ["private", "dm"]}, "members": user_id}
        ]
    }
    channels = await db.channels.find(query).to_list(100)
    return [str(ch["_id"]) for ch in channels]

@app.websocket("/ws/chat")
async def ws_chat(websocket: WebSocket):
    await websocket.accept()
    user = await ws_authenticate(websocket)
    if user is None:
        return

    user_id = user["id"]
    channel_ids = await get_user_channel_ids(user_id)

    for ch_id in channel_ids:
        await ws_manager.connect(ch_id, websocket)

    try:
        while True:
            data = await websocket.receive_json()
            msg_type = data.get("type")

            if msg_type == "message":
                channel_id = data.get("channel_id", "")
                content = data.get("content", "").strip()
                attachment = data.get("attachment")
                if not content and not attachment:
                    continue
                if len(content) > 2000:
                    continue
                if channel_id not in channel_ids:
                    continue

                now_iso = datetime.now(timezone.utc).isoformat()
                msg_doc = {
                    "channel_id": ObjectId(channel_id),
                    "sender_id": user_id,
                    "sender_name": user.get("full_name", ""),
                    "sender_role": user.get("role", ""),
                    "content": content,
                    "attachment": attachment,
                    "reactions": [],
                    "created_at": now_iso,
                    "edited_at": None,
                    "deleted": False,
                    "deleted_at": None,
                    "deleted_by": None
                }
                result = await db.messages.insert_one(msg_doc)
                msg_doc["id"] = str(result.inserted_id)
                msg_doc["channel_id"] = str(msg_doc["channel_id"])
                msg_doc.pop("_id", None)

                await ws_manager.broadcast(channel_id, {
                    "type": "new_message",
                    "message": msg_doc
                })

            elif msg_type == "typing":
                channel_id = data.get("channel_id", "")
                if channel_id not in channel_ids:
                    continue
                await ws_manager.broadcast(channel_id, {
                    "type": "typing",
                    "channel_id": channel_id,
                    "user_id": user_id,
                    "user_name": user.get("full_name", "")
                }, exclude=websocket)

            elif msg_type == "read":
                channel_id = data.get("channel_id", "")
                if channel_id not in channel_ids:
                    continue
                last_msg = await db.messages.find(
                    {"channel_id": ObjectId(channel_id), "deleted": False}
                ).sort("created_at", -1).limit(1).to_list(1)
                last_msg_id = last_msg[0]["_id"] if last_msg else None
                now_iso = datetime.now(timezone.utc).isoformat()
                await db.message_reads.update_one(
                    {"user_id": user_id, "channel_id": ObjectId(channel_id)},
                    {"$set": {
                        "last_read_message_id": last_msg_id,
                        "last_read_at": now_iso
                    }},
                    upsert=True
                )

    except WebSocketDisconnect:
        pass
    except Exception:
        pass
    finally:
        for ch_id in channel_ids:
            ws_manager.disconnect(ch_id, websocket)
